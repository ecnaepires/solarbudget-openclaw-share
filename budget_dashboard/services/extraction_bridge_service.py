import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, List, Optional
import re
import os
import sys
import tempfile
import traceback
import importlib
import io
import json

import pandas as pd
from openpyxl import load_workbook

from config import DIMENSIONAMENTO_CELL_MAP


def _normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.lower().split())


def _to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("R$", "").replace("r$", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def _count_pdf_pages_from_bytes(file_bytes: bytes) -> int:
    if not file_bytes:
        return 0
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            return int(len(pdf.pages))
    except Exception:
        return 0


def _attach_page_provenance(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()
    if "page_first_seen" not in out.columns:
        page_cols = [col for col in ["audit_header_page", "audit_itens_page", "audit_historico_page"] if col in out.columns]
        if page_cols:
            out["page_first_seen"] = (
                out[page_cols]
                .apply(pd.to_numeric, errors="coerce")
                .replace(0, pd.NA)
                .min(axis=1, skipna=True)
            )

    if "page_first_seen" not in out.columns:
        return out

    out["page_first_seen"] = pd.to_numeric(out["page_first_seen"], errors="coerce").fillna(0).astype(int)
    key_cols = [col for col in ["pdf_source", "uc"] if col in out.columns]
    if not key_cols:
        out["audit_pdf_pages"] = "[]"
        out.loc[out["page_first_seen"] > 0, "audit_pdf_pages"] = out.loc[
            out["page_first_seen"] > 0, "page_first_seen"
        ].apply(lambda val: json.dumps([int(val)], ensure_ascii=False))
        return out

    valid_pages = out[out["page_first_seen"] > 0]
    grouped_pages = valid_pages.groupby(key_cols)["page_first_seen"].apply(
        lambda series: sorted({int(value) for value in series.tolist() if int(value) > 0})
    )
    first_page_by_key = grouped_pages.apply(lambda values: int(values[0]) if values else 0).to_dict()
    pages_by_key = grouped_pages.apply(lambda values: json.dumps(values, ensure_ascii=False)).to_dict()

    def _row_key(row: pd.Series):
        if len(key_cols) == 1:
            return row[key_cols[0]]
        return tuple(row[col] for col in key_cols)

    out["page_first_seen"] = out.apply(
        lambda row: first_page_by_key.get(_row_key(row), int(row.get("page_first_seen", 0) or 0)),
        axis=1,
    )
    out["audit_pdf_pages"] = out.apply(
        lambda row: pages_by_key.get(_row_key(row), "[]"),
        axis=1,
    )
    return out


def default_extraction_root() -> Path:
    # Resolve relative to this file: services/ -> budget_dashboard/ -> repo root -> Project/
    repo_root = Path(__file__).resolve().parent.parent.parent
    candidate = (repo_root / "Project" / "estudo_faturas_municipios").resolve()
    if candidate.exists():
        return candidate
    return candidate


def scan_dimensionamento_sources(root_path: Path, limit: int = 200) -> List[dict]:
    root = Path(root_path)
    if not root.exists():
        return []

    candidates: List[Path] = []
    patterns = [
        "municipios/*/dimensionamento/*_dimensionamento.csv",
        "municipios/*/dimensionamento/*_dimensionamento.xlsx",
        "municipios/*/dimensionamento/*dimensionamento*.csv",
        "municipios/*/dimensionamento/*dimensionamento*.xlsx",
        "**/*filtered_flat*.xlsx",
        "**/*filtered_flat*.csv",
        "**/*sunergies_app*.xlsx",
        "**/*sunergies_app*.csv",
        "**/*_dimensionamento.csv",
        "**/*_dimensionamento.xlsx",
        "**/*dimensionamento*.csv",
        "**/*dimensionamento*.xlsx",
    ]
    for pattern in patterns:
        candidates.extend(root.glob(pattern))

    unique_files = {}
    for file_path in candidates:
        if file_path.is_file():
            unique_files[str(file_path.resolve())] = file_path.resolve()

    sorted_files = sorted(
        unique_files.values(),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )[:limit]

    sources: List[dict] = []
    for path in sorted_files:
        municipio = ""
        parts_upper = [part.upper() for part in path.parts]
        if "MUNICIPIOS" in parts_upper:
            idx = parts_upper.index("MUNICIPIOS")
            if len(path.parts) > idx + 1:
                municipio = path.parts[idx + 1]

        sources.append(
            {
                "path": str(path),
                "filename": path.name,
                "municipio": municipio,
                "modified_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                "ext": path.suffix.lower(),
                "kind": (
                    "streamlit_export"
                    if "filtered_flat" in path.name.lower() or "sunergies_app" in path.name.lower()
                    else "dimensionamento"
                ),
            }
        )

    return sources


def _load_dimensionamento_file(file_path: Path) -> pd.DataFrame:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, encoding="utf-8-sig")
    workbook = pd.ExcelFile(path)
    try:
        target_sheet = None
        for sheet_name in workbook.sheet_names:
            if "dimensionamento" in _normalize_text(sheet_name):
                target_sheet = sheet_name
                break
        if target_sheet is None and "displayed_data" in workbook.sheet_names:
            target_sheet = "displayed_data"
        if target_sheet is None:
            target_sheet = workbook.sheet_names[0]
        return pd.read_excel(workbook, sheet_name=target_sheet)
    finally:
        workbook.close()


def _find_column(columns: List[str], targets: List[str]) -> str | None:
    normalized = {column: _normalize_text(column) for column in columns}
    for target in targets:
        for column, norm in normalized.items():
            if target in norm:
                return column
    return None


def parse_dimensionamento_dataframe(df: pd.DataFrame, source_label: str = "") -> List[dict]:
    if df.empty:
        return []

    columns = list(df.columns)
    municipio_col = _find_column(columns, ["municipio"])
    cenario_col = _find_column(columns, ["cenario", "cenario"])
    mwp_col = _find_column(columns, ["potencia necessaria (mwp)", "potencia necessaria mwp"])
    kwp_col = _find_column(columns, ["potencia necessaria (kwp)", "potencia necessaria kwp"])
    prod_col = _find_column(columns, ["produtividade"])

    total_kwh_col = None
    for column in columns:
        norm = _normalize_text(column)
        if "total" in norm and "kwh" in norm:
            total_kwh_col = column
            break

    records: List[dict] = []
    for idx, row in df.iterrows():
        mwp_value = _to_float(row.get(mwp_col)) if mwp_col else None
        kwp_value = _to_float(row.get(kwp_col)) if kwp_col else None
        if mwp_value is None and kwp_value is not None:
            mwp_value = kwp_value / 1000.0
        if mwp_value is None:
            continue

        scenario_name = str(row.get(cenario_col, "")).strip() if cenario_col else ""
        if not scenario_name:
            scenario_name = f"Scenario Row {idx + 1}"

        records.append(
            {
                "municipio": str(row.get(municipio_col, "")).strip() if municipio_col else "",
                "cenario": scenario_name,
                "mwp": float(mwp_value),
                "kwp": float(kwp_value) if kwp_value is not None else float(mwp_value * 1000.0),
                "produtividade_kwh_kwp_ano": _to_float(row.get(prod_col)) if prod_col else None,
                "total_kwh": _to_float(row.get(total_kwh_col)) if total_kwh_col else None,
                "source_file": source_label,
            }
        )

    return records


def parse_dimensionamento_records(file_path: Path) -> List[dict]:
    df = _load_dimensionamento_file(file_path)
    records = parse_dimensionamento_dataframe(df, source_label=str(file_path))
    if records:
        return records
    return _parse_dimensionamento_template_workbook(Path(file_path))


def _sheet_cell_float(ws, cell_ref: str) -> float | None:
    try:
        value = ws[cell_ref].value
    except Exception:
        return None
    return _to_float(value)


def _parse_dimensionamento_template_workbook(file_path: Path) -> List[dict]:
    path = Path(file_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return []

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return []
    try:
        dim_sheet = next(
            (sheet for sheet in workbook.worksheets if "dimensionamento" in _normalize_text(sheet.title)),
            None,
        )
        if dim_sheet is None:
            return []

        _c = DIMENSIONAMENTO_CELL_MAP
        consumo_b3 = _sheet_cell_float(dim_sheet, _c["consumo_b3"]) or 0.0
        consumo_b4a = _sheet_cell_float(dim_sheet, _c["consumo_b4a"]) or 0.0
        consumo_a4_hp = _sheet_cell_float(dim_sheet, _c["consumo_a4_hp"]) or 0.0
        consumo_a4_fhp = _sheet_cell_float(dim_sheet, _c["consumo_a4_fhp"]) or 0.0

        kwp_b3 = _sheet_cell_float(dim_sheet, _c["kwp_b3"]) or 0.0
        kwp_b4a = _sheet_cell_float(dim_sheet, _c["kwp_b4a"]) or 0.0
        kwp_a4_hp = _sheet_cell_float(dim_sheet, _c["kwp_a4_hp"]) or 0.0
        kwp_a4_fhp = _sheet_cell_float(dim_sheet, _c["kwp_a4_fhp"]) or 0.0

        total_kwp = _sheet_cell_float(dim_sheet, _c["total_kwp"])
        if total_kwp is None:
            total_kwp = kwp_b3 + kwp_b4a + kwp_a4_hp + kwp_a4_fhp
        total_mwp = float(total_kwp) / 1000.0

        hsp = _sheet_cell_float(dim_sheet, _c["hsp"])
        performance_ratio = _sheet_cell_float(dim_sheet, _c["performance_ratio"])
        days_per_month = _sheet_cell_float(dim_sheet, _c["days_per_month"])
        a4_hp_factor = _sheet_cell_float(dim_sheet, _c["a4_hp_factor"])
        capex_brl_per_mwp = _sheet_cell_float(dim_sheet, _c["capex_brl_per_mwp"])

        investment_brl = _sheet_cell_float(dim_sheet, _c["investment_brl"])
        if investment_brl is None and capex_brl_per_mwp is not None:
            investment_brl = total_mwp * float(capex_brl_per_mwp)
        if investment_brl is None:
            investment_brl = 0.0

        energy_cost_month = _sheet_cell_float(dim_sheet, _c["energy_cost_month"])
        payback_months = _sheet_cell_float(dim_sheet, _c["payback_months"])

        monthly_energy_cost_brl: dict[str, float | None] = {"TOTAL": energy_cost_month}
        tariff_map = {"B3": None, "B4A": None, "A4_HP": None, "A4_FHP": None}
        payback_needs_tariff_input = payback_months is None and energy_cost_month is None

        return [
            {
                "municipio": _infer_municipio(path.name),
                "cenario": "Dimensionamento Previo UFV (Planilha)",
                "mwp": float(total_mwp),
                "kwp": float(total_kwp),
                "total_kwh": float(consumo_b3 + consumo_b4a + consumo_a4_hp + consumo_a4_fhp),
                "source_file": str(path),
                "dimensionamento_inputs": {
                    "hsp": hsp,
                    "performance_ratio": performance_ratio,
                    "days_per_month": days_per_month,
                    "a4_hp_factor": a4_hp_factor,
                    "capex_brl_per_mwp": capex_brl_per_mwp,
                },
                "consumo_medio_mensal_kwh": {
                    "B3": float(consumo_b3),
                    "B4A": float(consumo_b4a),
                    "A4_HP": float(consumo_a4_hp),
                    "A4_FHP": float(consumo_a4_fhp),
                },
                "ufv_kwp_por_categoria": {
                    "B3": float(kwp_b3),
                    "B4A": float(kwp_b4a),
                    "A4_HP": float(kwp_a4_hp),
                    "A4_FHP": float(kwp_a4_fhp),
                },
                "ufv_total_kwp": float(total_kwp),
                "ufv_total_mwp": float(total_mwp),
                "investment_brl": float(investment_brl),
                "tariff_rs_kwh": tariff_map,
                "monthly_energy_cost_brl": monthly_energy_cost_brl,
                "payback_months": float(payback_months) if payback_months is not None else None,
                "payback_needs_tariff_input": payback_needs_tariff_input,
                "warnings": ["Imported from pre-calculated Dimensionamento worksheet cells."],
            }
        ]
    finally:
        try:
            workbook.close()
        except Exception:
            pass


MONTH_MAP = {
    "JAN": "01",
    "FEV": "02",
    "MAR": "03",
    "ABR": "04",
    "MAI": "05",
    "JUN": "06",
    "JUL": "07",
    "AGO": "08",
    "SET": "09",
    "OUT": "10",
    "NOV": "11",
    "DEZ": "12",
}


def _parse_reference_to_date(value) -> pd.Timestamp:
    if value is None:
        return pd.NaT
    text = str(value).strip()
    if not text:
        return pd.NaT

    mm_yyyy = pd.to_datetime(text, format="%m/%Y", errors="coerce")
    if pd.notna(mm_yyyy):
        return mm_yyyy

    match = re.match(r"^(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)[/-](\d{2,4})$", text.upper())
    if match:
        month = MONTH_MAP[match.group(1)]
        year = match.group(2)
        if len(year) == 2:
            year = f"20{year}"
        return pd.to_datetime(f"{year}-{month}-01", errors="coerce")

    return pd.to_datetime(text, errors="coerce")


def parse_streamlit_export_records(
    file_path: Path,
    months_to_use: int = 13,
    hsp: float = 4.9,
    performance_ratio: float = 0.80,
    days_per_month: float = 30.0,
    a4_hp_factor: float = 1.0,
    capex_brl_per_mwp: float = 8_500_000.0,
    tariff_b3_rs_kwh: float | None = None,
    tariff_b4a_rs_kwh: float | None = None,
    tariff_a4_hp_rs_kwh: float | None = None,
    tariff_a4_fhp_rs_kwh: float | None = None,
) -> List[dict]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path, encoding="utf-8-sig")
    else:
        workbook = pd.ExcelFile(path)
        target_sheet = "displayed_data" if "displayed_data" in workbook.sheet_names else workbook.sheet_names[0]
        df = pd.read_excel(workbook, sheet_name=target_sheet)

    return parse_streamlit_export_dataframe(
        df=df,
        source_label=str(path),
        months_to_use=months_to_use,
        hsp=hsp,
        performance_ratio=performance_ratio,
        days_per_month=days_per_month,
        a4_hp_factor=a4_hp_factor,
        capex_brl_per_mwp=capex_brl_per_mwp,
        tariff_b3_rs_kwh=tariff_b3_rs_kwh,
        tariff_b4a_rs_kwh=tariff_b4a_rs_kwh,
        tariff_a4_hp_rs_kwh=tariff_a4_hp_rs_kwh,
        tariff_a4_fhp_rs_kwh=tariff_a4_fhp_rs_kwh,
    )


def _find_col_by_aliases(columns: List[str], aliases: List[str]) -> str | None:
    normalized = {col: _normalize_text(col) for col in columns}
    targets = [_normalize_text(alias) for alias in aliases]

    for target in targets:
        for col, norm in normalized.items():
            if norm == target:
                return col
    for target in targets:
        for col, norm in normalized.items():
            if target and target in norm:
                return col
    return None


def _coerce_numeric_col(df: pd.DataFrame, column_name: str | None) -> pd.Series:
    if column_name is None or column_name not in df.columns:
        return pd.Series(0.0, index=df.index, dtype="float64")
    return pd.to_numeric(df[column_name], errors="coerce").fillna(0.0)


def _normalize_consumer_class(value) -> str:
    text = _normalize_text(value)
    if not text:
        return "OUTROS"
    if "a4" in text or "alta tensao" in text or "grupo a" in text or text == "a":
        return "A4"
    if text == "ip" or "iluminacao publica" in text or "iluminacao" in text or "b4a" in text or "b4" in text:
        return "IP"
    if "b3" in text or "grupo b" in text or "comercial" in text:
        return "B3"
    return "OUTROS"


def _to_month_start(ts: pd.Timestamp) -> pd.Timestamp:
    return pd.Timestamp(year=ts.year, month=ts.month, day=1)


def _expected_month_window(max_reference: pd.Timestamp, months_to_use: int) -> list[pd.Timestamp]:
    months = max(1, int(months_to_use))
    current = _to_month_start(max_reference)
    refs: list[pd.Timestamp] = []
    for _ in range(months):
        refs.append(current)
        current = current - pd.DateOffset(months=1)
    return sorted(refs)


def _monthly_series_for_class(
    group: pd.DataFrame,
    expected_months: list[pd.Timestamp],
    consumer_class: str,
    value_col: str,
) -> pd.Series:
    if not expected_months:
        return pd.Series(dtype="float64")

    class_df = group[group["_consumer_class_bridge"] == consumer_class]
    if class_df.empty or value_col not in class_df.columns:
        return pd.Series(0.0, index=expected_months, dtype="float64")

    monthly = (
        class_df.groupby("_reference_month_bridge", dropna=False)[value_col]
        .sum()
        .reindex(expected_months, fill_value=0.0)
    )
    return monthly.astype(float)


def _optional_positive_float(value) -> float | None:
    parsed = _to_float(value)
    if parsed is None:
        return None
    if parsed <= 0:
        return None
    return float(parsed)


def parse_streamlit_export_dataframe(
    df: pd.DataFrame,
    source_label: str = "",
    months_to_use: int = 13,
    hsp: float = 4.9,
    performance_ratio: float = 0.80,
    days_per_month: float = 30.0,
    a4_hp_factor: float = 1.0,
    capex_brl_per_mwp: float = 8_500_000.0,
    tariff_b3_rs_kwh: float | None = None,
    tariff_b4a_rs_kwh: float | None = None,
    tariff_a4_hp_rs_kwh: float | None = None,
    tariff_a4_fhp_rs_kwh: float | None = None,
) -> List[dict]:
    if df.empty:
        return []

    if int(months_to_use) <= 0:
        raise ValueError("months_to_use must be greater than 0.")
    if float(hsp) <= 0:
        raise ValueError("HSP must be greater than 0.")
    if float(performance_ratio) <= 0:
        raise ValueError("Performance ratio (PR) must be greater than 0.")
    if float(days_per_month) <= 0:
        raise ValueError("Days per month must be greater than 0.")
    if float(a4_hp_factor) <= 0:
        raise ValueError("A4 HP factor must be greater than 0.")
    if float(capex_brl_per_mwp) < 0:
        raise ValueError("CAPEX BRL/MWp cannot be negative.")

    local = df.copy()
    columns = list(local.columns)

    municipality_col = _find_col_by_aliases(columns, ["municipio", "municipality", "cidade"])
    uc_col = _find_col_by_aliases(columns, ["uc", "unidade consumidora", "unidade_consumidora", "instalacao"])
    if uc_col is None:
        uc_col = "__uc_fallback__"
        local[uc_col] = local.index.astype(str)

    date_col = _find_col_by_aliases(
        columns,
        ["reference_date", "data_referencia", "referencia", "competencia", "mes referencia"],
    )
    if date_col is None:
        return []

    class_col = _find_col_by_aliases(
        columns,
        [
            "categoria",
            "consumer_class",
            "classificacao_uc",
            "classificacao (categoria)",
            "grupo tarifario",
            "subgrupo",
            "tarifa",
            "tariff",
            "classe",
        ],
    )
    b3_ip_col = _find_col_by_aliases(
        columns,
        [
            "kwh_b3_ip",
            "consumo te (kwh)",
            "consumo te",
            "itens_fatura_energia_kwh",
            "kwh_total_te",
            "kwh",
        ],
    )
    a4_hp_col = _find_col_by_aliases(
        columns,
        [
            "consumo_hp_kwh",
            "kwh_a4_p_te",
            "consumo te ponta",
            "consumo ponta",
            "consumo hp",
            "consumo hora ponta",
        ],
    )
    a4_fhp_col = _find_col_by_aliases(
        columns,
        [
            "consumo_fhp_kwh",
            "kwh_a4_fp_te",
            "consumo te fora ponta",
            "consumo fora ponta",
            "consumo fhp",
            "consumo hfp",
            "consumo unico",
        ],
    )
    a4_total_col = _find_col_by_aliases(columns, ["kwh_total_te", "itens_fatura_energia_kwh", "kwh"])

    local["_municipio_bridge"] = (
        local[municipality_col].astype(str).str.strip() if municipality_col else "NAO INFORMADO"
    )
    local["_uc_bridge"] = local[uc_col].astype(str).str.strip()
    local["_reference_bridge"] = local[date_col].apply(_parse_reference_to_date)
    local = local[local["_reference_bridge"].notna()].copy()
    if local.empty:
        return []

    local["_reference_month_bridge"] = local["_reference_bridge"].apply(_to_month_start)
    local["_b3_ip_kwh_bridge"] = _coerce_numeric_col(local, b3_ip_col)
    local["_a4_hp_kwh_bridge"] = _coerce_numeric_col(local, a4_hp_col)
    local["_a4_fhp_kwh_bridge"] = _coerce_numeric_col(local, a4_fhp_col)
    local["_a4_total_kwh_bridge"] = _coerce_numeric_col(local, a4_total_col)
    local["_consumer_class_bridge"] = (
        local[class_col].apply(_normalize_consumer_class) if class_col else "OUTROS"
    )

    hp_has_source = a4_hp_col is not None
    fhp_has_source = a4_fhp_col is not None

    tariff_b3 = _optional_positive_float(tariff_b3_rs_kwh)
    tariff_b4a = _optional_positive_float(tariff_b4a_rs_kwh)
    tariff_a4_hp = _optional_positive_float(tariff_a4_hp_rs_kwh)
    tariff_a4_fhp = _optional_positive_float(tariff_a4_fhp_rs_kwh)

    generation_kwh_per_kwp_month = float(hsp) * float(performance_ratio) * float(days_per_month)
    records: List[dict] = []

    for municipio, group in local.groupby("_municipio_bridge", dropna=False):
        max_reference = group["_reference_bridge"].max()
        if pd.isna(max_reference):
            continue

        selected_months = _expected_month_window(max_reference, int(months_to_use))
        month_labels = [month.strftime("%m/%Y") for month in selected_months]

        b3_monthly = _monthly_series_for_class(group, selected_months, "B3", "_b3_ip_kwh_bridge")
        b4a_monthly = _monthly_series_for_class(group, selected_months, "IP", "_b3_ip_kwh_bridge")
        a4_hp_monthly = _monthly_series_for_class(group, selected_months, "A4", "_a4_hp_kwh_bridge")
        a4_fhp_monthly = _monthly_series_for_class(group, selected_months, "A4", "_a4_fhp_kwh_bridge")
        if not hp_has_source and not fhp_has_source:
            a4_fhp_monthly = _monthly_series_for_class(group, selected_months, "A4", "_a4_total_kwh_bridge")

        consumo_b3 = float(b3_monthly.mean()) if not b3_monthly.empty else 0.0
        consumo_b4a = float(b4a_monthly.mean()) if not b4a_monthly.empty else 0.0
        consumo_a4_hp = float(a4_hp_monthly.mean()) if not a4_hp_monthly.empty else 0.0
        consumo_a4_fhp = float(a4_fhp_monthly.mean()) if not a4_fhp_monthly.empty else 0.0

        kwp_b3 = consumo_b3 / generation_kwh_per_kwp_month
        kwp_b4a = consumo_b4a / generation_kwh_per_kwp_month
        kwp_a4_hp = (consumo_a4_hp / generation_kwh_per_kwp_month) / float(a4_hp_factor)
        kwp_a4_fhp = consumo_a4_fhp / generation_kwh_per_kwp_month
        total_kwp = kwp_b3 + kwp_b4a + kwp_a4_hp + kwp_a4_fhp
        total_mwp = total_kwp / 1000.0
        investment_brl = total_mwp * float(capex_brl_per_mwp)

        consumo_map = {
            "B3": consumo_b3,
            "B4A": consumo_b4a,
            "A4_HP": consumo_a4_hp,
            "A4_FHP": consumo_a4_fhp,
        }
        tariffs_map = {
            "B3": tariff_b3,
            "B4A": tariff_b4a,
            "A4_HP": tariff_a4_hp,
            "A4_FHP": tariff_a4_fhp,
        }

        monthly_cost_map: dict[str, float | None] = {}
        payback_ready = True
        total_energy_cost_month_brl = 0.0
        for key, consumo_medio in consumo_map.items():
            tariff_value = tariffs_map[key]
            if consumo_medio <= 0:
                monthly_cost_map[key] = 0.0
                continue
            if tariff_value is None:
                monthly_cost_map[key] = None
                payback_ready = False
                continue
            monthly_cost = consumo_medio * float(tariff_value)
            monthly_cost_map[key] = monthly_cost
            total_energy_cost_month_brl += monthly_cost

        monthly_cost_map["TOTAL"] = total_energy_cost_month_brl if payback_ready else None
        payback_months = (
            investment_brl / total_energy_cost_month_brl
            if payback_ready and total_energy_cost_month_brl > 0
            else None
        )

        records.append(
            {
                "municipio": str(municipio).strip(),
                "cenario": "Dimensionamento Previo UFV",
                "mwp": float(total_mwp),
                "kwp": float(total_kwp),
                "total_kwh": float(sum(consumo_map.values()) * len(selected_months)),
                "source_file": source_label,
                "months_to_use": int(months_to_use),
                "month_labels": month_labels,
                "dimensionamento_inputs": {
                    "hsp": float(hsp),
                    "performance_ratio": float(performance_ratio),
                    "days_per_month": float(days_per_month),
                    "a4_hp_factor": float(a4_hp_factor),
                    "capex_brl_per_mwp": float(capex_brl_per_mwp),
                },
                "consumo_medio_mensal_kwh": consumo_map,
                "consumo_mensal_janela_kwh": {
                    "B3": [float(v) for v in b3_monthly.tolist()],
                    "B4A": [float(v) for v in b4a_monthly.tolist()],
                    "A4_HP": [float(v) for v in a4_hp_monthly.tolist()],
                    "A4_FHP": [float(v) for v in a4_fhp_monthly.tolist()],
                },
                "ufv_kwp_por_categoria": {
                    "B3": float(kwp_b3),
                    "B4A": float(kwp_b4a),
                    "A4_HP": float(kwp_a4_hp),
                    "A4_FHP": float(kwp_a4_fhp),
                },
                "ufv_total_kwp": float(total_kwp),
                "ufv_total_mwp": float(total_mwp),
                "generation_kwh_per_kwp_month": float(generation_kwh_per_kwp_month),
                "investment_brl": float(investment_brl),
                "tariff_rs_kwh": {
                    "B3": tariff_b3,
                    "B4A": tariff_b4a,
                    "A4_HP": tariff_a4_hp,
                    "A4_FHP": tariff_a4_fhp,
                },
                "monthly_energy_cost_brl": monthly_cost_map,
                "payback_months": float(payback_months) if payback_months is not None else None,
                "payback_needs_tariff_input": not payback_ready,
                "warnings": (
                    ["A4 HP/FHP columns not found. Falling back to A4 total kWh as FHP."]
                    if (not hp_has_source and not fhp_has_source)
                    else []
                ),
            }
        )

    return records


def _infer_municipio(file_name: str, municipio_override: str = "") -> str:
    override = str(municipio_override or "").strip()
    if override:
        return override

    stem = Path(file_name).stem
    stem = re.sub(r"(?i)[_-]?faturas?.*$", "", stem)
    stem = re.sub(r"(?i)[_-]?invoice[s]?.*$", "", stem)
    stem = re.sub(r"[_\-]+", " ", stem).strip()
    return stem or "NAO INFORMADO"


def _get_extract_pdf_callable(extraction_root: Path):
    root = Path(extraction_root).resolve()
    if not root.exists():
        raise FileNotFoundError(f"Extraction project folder not found: {root}")
    package_init = root / "fatura_engine" / "__init__.py"
    extractors_file = root / "fatura_engine" / "extractors.py"
    if not package_init.exists() or not extractors_file.exists():
        raise FileNotFoundError(
            "Expected fatura_engine package not found in extraction root. "
            f"Missing: {package_init if not package_init.exists() else extractors_file}"
        )

    root_str = str(root)
    if root_str not in sys.path:
        sys.path.insert(0, root_str)

    try:
        def _is_under_root(module_file: str, expected_root: Path) -> bool:
            try:
                module_path = Path(module_file).resolve()
            except Exception:
                return False
            try:
                return module_path.is_relative_to(expected_root)
            except AttributeError:
                return str(module_path).lower().startswith(str(expected_root).lower())

        # If fatura_engine was already imported from another location, clear it.
        imported_pkg = sys.modules.get("fatura_engine")
        if imported_pkg is not None:
            imported_file = getattr(imported_pkg, "__file__", "")
            if not _is_under_root(str(imported_file), root):
                for key in list(sys.modules.keys()):
                    if key == "fatura_engine" or key.startswith("fatura_engine."):
                        sys.modules.pop(key, None)

        package_module = importlib.import_module("fatura_engine")
        package_file = getattr(package_module, "__file__", "")
        if not _is_under_root(str(package_file), root):
            # Force reload from selected root.
            for key in list(sys.modules.keys()):
                if key == "fatura_engine" or key.startswith("fatura_engine."):
                    sys.modules.pop(key, None)
            package_module = importlib.import_module("fatura_engine")

        extractors_module = importlib.import_module("fatura_engine.extractors")
        extractors_file_loaded = getattr(extractors_module, "__file__", "")
        if not _is_under_root(str(extractors_file_loaded), root):
            for key in list(sys.modules.keys()):
                if key == "fatura_engine" or key.startswith("fatura_engine."):
                    sys.modules.pop(key, None)
            extractors_module = importlib.import_module("fatura_engine.extractors")

        extract_pdf = getattr(extractors_module, "extract_pdf", None)
        if not callable(extract_pdf):
            raise AttributeError("fatura_engine.extractors.extract_pdf was not found/callable")
    except Exception as exc:
        details = "".join(traceback.format_exception_only(type(exc), exc)).strip()
        raise ImportError(
            "Could not import fatura_engine.extractors.extract_pdf from extraction project. "
            f"Root: {root}. Details: {details}"
        ) from exc

    return extract_pdf


def extract_records_from_uploaded_pdfs(
    uploaded_pdfs,
    extraction_root: Path,
    municipio_override: str = "",
    expand_a4_historico: bool = True,
    months_to_use: int = 13,
    hsp: float = 4.9,
    performance_ratio: float = 0.80,
    days_per_month: float = 30.0,
    a4_hp_factor: float = 1.0,
    capex_brl_per_mwp: float = 8_500_000.0,
    tariff_b3_rs_kwh: float | None = None,
    tariff_b4a_rs_kwh: float | None = None,
    tariff_a4_hp_rs_kwh: float | None = None,
    tariff_a4_fhp_rs_kwh: float | None = None,
    progress_callback: Optional[Callable[[dict[str, Any]], None]] = None,
    group_batch_when_municipio_empty: bool = True,
):
    extract_pdf = _get_extract_pdf_callable(extraction_root)

    def _emit_progress(event: dict[str, Any]) -> None:
        if not progress_callback:
            return
        try:
            progress_callback(event)
        except Exception:
            # Progress UI must never break extraction.
            pass

    frames: List[pd.DataFrame] = []
    files_processed = 0
    failed_files: List[str] = []
    discovery_reports: dict[str, dict[str, Any]] = {}
    pdf_jobs: List[dict] = []
    total_pages = 0
    municipio_override_clean = str(municipio_override or "").strip()
    batch_municipio_fallback = municipio_override_clean
    if not batch_municipio_fallback and bool(group_batch_when_municipio_empty) and len(uploaded_pdfs) > 1:
        batch_municipio_fallback = "MUNICIPIO_NAO_INFORMADO"

    for uploaded in uploaded_pdfs:
        file_name = getattr(uploaded, "name", "uploaded.pdf")
        try:
            file_bytes = uploaded.getvalue()
        except Exception:
            failed_files.append(file_name)
            continue

        page_count = _count_pdf_pages_from_bytes(file_bytes)
        total_pages += page_count
        pdf_jobs.append(
            {
                "name": file_name,
                "bytes": file_bytes,
                "pages": page_count,
            }
        )

    pages_processed = 0

    _emit_progress(
            {
                "phase": "init",
                "pages_processed": pages_processed,
                "total_pages": total_pages,
                "progress": 0.0,
                "current_pdf_name": "",
                "pdf_page_no": 0,
                "pdf_pages": 0,
            }
    )

    previous_discovery_output_dir = os.getenv("DISCOVERY_OUTPUT_DIR")
    try:
        with tempfile.TemporaryDirectory(prefix="pdf_discovery_") as discovery_dir:
            os.environ["DISCOVERY_OUTPUT_DIR"] = discovery_dir

            for job in pdf_jobs:
                file_name = job["name"]
                file_bytes = job["bytes"]
                pdf_pages = int(job.get("pages", 0) or 0)

                if total_pages > 0 and pdf_pages > 0:
                    _emit_progress(
                        {
                            "phase": "page_start",
                            "pages_processed": pages_processed,
                            "total_pages": total_pages,
                            "progress": pages_processed / total_pages,
                            "current_pdf_name": file_name,
                            "pdf_page_no": 1,
                            "pdf_pages": pdf_pages,
                        }
                    )

                tmp_path = ""
                current_pdf_processed = 0
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                        tmp_file.write(file_bytes)
                        tmp_path = tmp_file.name

                    def _on_pdf_page_done(local_page_no: int, local_total_pages: int) -> None:
                        nonlocal pages_processed, current_pdf_processed, total_pages, pdf_pages

                        if pdf_pages <= 0 and int(local_total_pages) > 0:
                            pdf_pages = int(local_total_pages)
                            total_pages += int(local_total_pages)

                        current_pdf_processed = int(local_page_no)
                        pages_processed = min(total_pages, pages_processed + 1) if total_pages > 0 else pages_processed + 1
                        progress = (pages_processed / total_pages) if total_pages > 0 else 1.0

                        _emit_progress(
                            {
                                "phase": "page_done",
                                "pages_processed": pages_processed,
                                "total_pages": total_pages,
                                "progress": progress,
                                "current_pdf_name": file_name,
                                "pdf_page_no": int(local_page_no),
                                "pdf_pages": int(local_total_pages),
                            }
                        )

                        if int(local_page_no) < int(local_total_pages):
                            _emit_progress(
                                {
                                    "phase": "page_start",
                                    "pages_processed": pages_processed,
                                    "total_pages": total_pages,
                                    "progress": progress,
                                    "current_pdf_name": file_name,
                                    "pdf_page_no": int(local_page_no) + 1,
                                    "pdf_pages": int(local_total_pages),
                                }
                            )

                    df = extract_pdf(
                        tmp_path,
                        expand_a4_historico=bool(expand_a4_historico),
                        progress_callback=_on_pdf_page_done if progress_callback else None,
                        discovery_mode=True,
                    )
                    files_processed += 1

                    discovery_summary = {}
                    if isinstance(df, pd.DataFrame):
                        discovery_summary = dict(df.attrs.get("discovery_summary") or {})
                    if discovery_summary:
                        discovery_summary["pdf_source"] = file_name
                        discovery_reports[file_name] = discovery_summary

                    if df is None or df.empty:
                        continue

                    local_df = df.copy()
                    local_df["uploaded_file"] = file_name
                    inferred_municipio = _infer_municipio(file_name, batch_municipio_fallback)
                    if "municipio" not in local_df.columns:
                        local_df["municipio"] = inferred_municipio
                    else:
                        local_df["municipio"] = (
                            local_df["municipio"]
                            .fillna("")
                            .astype(str)
                            .str.strip()
                            .replace("", inferred_municipio)
                        )

                    frames.append(local_df)
                except Exception as exc:
                    failed_files.append(file_name)
                    discovery_reports[file_name] = {
                        "pdf_source": file_name,
                        "detected_layout": "",
                        "effective_layout": "",
                        "field_extraction_report": {},
                        "unmatched_text_sample": "",
                        "error": f"{type(exc).__name__}: {exc}",
                    }
                finally:
                    if total_pages > 0 and pdf_pages > current_pdf_processed:
                        remaining = int(pdf_pages - current_pdf_processed)
                        pages_processed = min(total_pages, pages_processed + remaining)
                        _emit_progress(
                            {
                                "phase": "pdf_skipped",
                                "pages_processed": pages_processed,
                                "total_pages": total_pages,
                                "progress": pages_processed / total_pages,
                                "current_pdf_name": file_name,
                                "pdf_page_no": int(pdf_pages),
                                "pdf_pages": int(pdf_pages),
                            }
                        )

                    if tmp_path and os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except OSError:
                            pass

    finally:
        if previous_discovery_output_dir is None:
            os.environ.pop("DISCOVERY_OUTPUT_DIR", None)
        else:
            os.environ["DISCOVERY_OUTPUT_DIR"] = previous_discovery_output_dir

    if not frames:
        return [], {
            "files_processed": files_processed,
            "rows_extracted": 0,
            "failed_files": failed_files,
            "total_pages": total_pages,
            "pages_processed": pages_processed,
            "discovery_reports": discovery_reports,
            "master_df": pd.DataFrame(),
        }

    master_df = pd.concat(frames, ignore_index=True)
    records = parse_streamlit_export_dataframe(
        master_df,
        source_label=f"PDF upload ({len(uploaded_pdfs)} file(s))",
        months_to_use=months_to_use,
        hsp=hsp,
        performance_ratio=performance_ratio,
        days_per_month=days_per_month,
        a4_hp_factor=a4_hp_factor,
        capex_brl_per_mwp=capex_brl_per_mwp,
        tariff_b3_rs_kwh=tariff_b3_rs_kwh,
        tariff_b4a_rs_kwh=tariff_b4a_rs_kwh,
        tariff_a4_hp_rs_kwh=tariff_a4_hp_rs_kwh,
        tariff_a4_fhp_rs_kwh=tariff_a4_fhp_rs_kwh,
    )
    master_df_with_page_provenance = _attach_page_provenance(master_df)

    return records, {
        "files_processed": files_processed,
        "rows_extracted": int(len(master_df)),
        "failed_files": failed_files,
        "total_pages": total_pages,
        "pages_processed": pages_processed,
        "discovery_reports": discovery_reports,
        "master_df": master_df_with_page_provenance,
    }
