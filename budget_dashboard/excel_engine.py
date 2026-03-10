import re
import unicodedata
from numbers import Number
from pathlib import Path

from openpyxl import load_workbook

from config import OUTPUT_PATH, TEMPLATE_PATH

CURRENCY_FORMAT_PT_BR = "[$R$-pt-BR] #,##0.00"
DYNAMIC_TOTAL_FIXES = [
    ("B3", "G410", "G", 5),
    ("B4A Ilum. Pública", "G47", "G", 5),
    ("A4 VERDE", "E18", "E", 5),
    ("A4 VERDE", "G18", "G", 5),
    ("ANÁLISE CONSUMO A4 VERDE X ACL", "AE36", "AE", 5),
]


def _normalize_sheet_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.upper().split())


def _resolve_sheet_name(workbook, expected_name: str) -> str | None:
    if expected_name in workbook.sheetnames:
        return expected_name
    target = _normalize_sheet_name(expected_name)
    for sheet_name in workbook.sheetnames:
        if _normalize_sheet_name(sheet_name) == target:
            return sheet_name
    return None


def _is_number(value) -> bool:
    if isinstance(value, bool):
        return False
    return isinstance(value, Number)


def _find_last_numeric_row(ws, col: str, start_row: int) -> int | None:
    last = None
    for row_idx in range(int(start_row), ws.max_row + 1):
        if _is_number(ws[f"{col}{row_idx}"].value):
            last = row_idx
    return last


def _extract_row_number(cell_ref: str) -> int | None:
    match = re.search(r"(\d+)$", str(cell_ref or ""))
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _build_dynamic_sum_formula(total_cell: str, col: str, start_row: int, end_row: int) -> str:
    total_row = _extract_row_number(total_cell)
    if total_row is None:
        return f"=SUM({col}{start_row}:{col}{end_row})"
    if total_row < start_row or total_row > end_row:
        return f"=SUM({col}{start_row}:{col}{end_row})"

    left_end = total_row - 1
    right_start = total_row + 1
    parts = []
    if left_end >= start_row:
        parts.append(f"SUM({col}{start_row}:{col}{left_end})")
    if right_start <= end_row:
        parts.append(f"SUM({col}{right_start}:{col}{end_row})")
    if not parts:
        return "=0"
    if len(parts) == 1:
        return f"={parts[0]}"
    return f"={parts[0]}+{parts[1]}"


def apply_dynamic_totals(workbook, total_fixes=None) -> None:
    fixes = list(total_fixes or DYNAMIC_TOTAL_FIXES)
    for sheet_name, total_cell, col, start_row in fixes:
        resolved_name = _resolve_sheet_name(workbook, sheet_name)
        if not resolved_name:
            continue

        ws = workbook[resolved_name]
        last_row = _find_last_numeric_row(ws, col, start_row)
        if last_row is None or last_row < start_row:
            ws[total_cell].value = f"=SUM({col}{start_row}:{col}{start_row})"
            continue

        ws[total_cell].value = _build_dynamic_sum_formula(
            total_cell=total_cell,
            col=col,
            start_row=start_row,
            end_row=last_row,
        )


def _resolve_sheet_by_prefix(workbook, prefix: str):
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith(prefix):
            return workbook[sheet_name]
    raise KeyError(f"No sheet found for prefix '{prefix}'")


def write_excel(
    total_kwp,
    module_data,
    inverter_qty,
    inverter_data,
    output_path: Path = OUTPUT_PATH,
    template_path: Path = TEMPLATE_PATH,
):
    """Write core PV values into the template and save workbook."""
    wb = load_workbook(template_path, data_only=False)

    ws_mod = _resolve_sheet_by_prefix(wb, "7.1")
    ws_mod["E5"].value = total_kwp
    ws_mod["F5"].value = module_data["price_sem_bdi_per_kwp"]
    ws_mod["G5"].value = module_data["price_com_bdi_per_kwp"]
    ws_mod["F5"].number_format = CURRENCY_FORMAT_PT_BR
    ws_mod["G5"].number_format = CURRENCY_FORMAT_PT_BR

    ws_inv = _resolve_sheet_by_prefix(wb, "7.2")
    ws_inv["E5"].value = inverter_qty
    ws_inv["F5"].value = inverter_data["price_sem_bdi"]
    ws_inv["G5"].value = inverter_data["price_com_bdi"]
    ws_inv["F5"].number_format = CURRENCY_FORMAT_PT_BR
    ws_inv["G5"].number_format = CURRENCY_FORMAT_PT_BR

    apply_dynamic_totals(wb)
    wb.save(output_path)


def write_full_budget_excel(
    updates,
    output_path: Path = OUTPUT_PATH,
    template_path: Path = TEMPLATE_PATH,
):
    """Write quantity and unit prices to mapped cells and save workbook."""
    wb = load_workbook(template_path, data_only=False)

    for update in updates:
        ws = wb[update["sheet_name"]]
        if update.get("quantity") is not None:
            ws[update["quantity_cell"]].value = update["quantity"]
        if update.get("price_sem") is not None:
            ws[update["price_sem_cell"]].value = update["price_sem"]
            ws[update["price_sem_cell"]].number_format = CURRENCY_FORMAT_PT_BR
        if update.get("price_com") is not None:
            ws[update["price_com_cell"]].value = update["price_com"]
            ws[update["price_com_cell"]].number_format = CURRENCY_FORMAT_PT_BR

    apply_dynamic_totals(wb)
    wb.save(output_path)
