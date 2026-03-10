import argparse
import os
import re
import shutil
import time
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from numbers import Number

import pandas as pd


from fatura_engine.extractors import extract_pdf
from fatura_engine.audit import build_audit_pdf_pages
from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator

MONTHS_TO_USE = 13  # default: most recent 13 months
GLOBAL_TEMPLATE_CANDIDATES = [
    os.path.join("templates", "final_template_all_in_one.xlsx"),
    "final_template_all_in_one.xlsx",
    "PALHOÇA_PROJETO_final_model_filledCORRECTVERSION.xlsx",
]

DYNAMIC_TOTAL_FIXES = [
    ("B3", "G410", "G", 5),
    ("B4A Ilum. Pública", "G47", "G", 5),
    ("A4 VERDE", "E18", "E", 5),
    ("A4 VERDE", "G18", "G", 5),
    ("ANÁLISE CONSUMO A4 VERDE X ACL", "AE36", "AE", 5),
]

def _normalize_text(s: str) -> str:
    txt = unicodedata.normalize("NFKD", str(s))
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    return " ".join(txt.lower().split())


def _sheet_exists_like(sheetnames: list[str], target: str) -> bool:
    t = _normalize_text(target)
    for name in sheetnames:
        n = _normalize_text(name)
        if n == t or t in n:
            return True
    return False



def _normalize_sheet_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.upper().split())


def _resolve_sheet_name(workbook, expected_name: str) -> str | None:
    if expected_name in workbook.sheetnames:
        return expected_name
    target = _normalize_sheet_name(expected_name)
    for sheet_name in workbook.sheetnames:
        if _normalize_sheet_name(sheet_name) == target:
            return sheet_name
    return None


def _is_number(value) -> bool:
    if isinstance(value, bool):
        return False
    return isinstance(value, Number)


def _find_last_numeric_row(ws, col: str, start_row: int) -> int | None:
    last = None
    for row_idx in range(int(start_row), ws.max_row + 1):
        if _is_number(ws[f"{col}{row_idx}"].value):
            last = row_idx
    return last


def _extract_row_number(cell_ref: str) -> int | None:
    match = re.search(r"(\d+)$", str(cell_ref or ""))
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _build_dynamic_sum_formula(total_cell: str, col: str, start_row: int, end_row: int) -> str:
    total_row = _extract_row_number(total_cell)
    if total_row is None:
        return f"=SUM({col}{start_row}:{col}{end_row})"
    if total_row < start_row or total_row > end_row:
        return f"=SUM({col}{start_row}:{col}{end_row})"

    left_end = total_row - 1
    right_start = total_row + 1
    parts = []
    if left_end >= start_row:
        parts.append(f"SUM({col}{start_row}:{col}{left_end})")
    if right_start <= end_row:
        parts.append(f"SUM({col}{right_start}:{col}{end_row})")
    if not parts:
        return "=0"
    if len(parts) == 1:
        return f"={parts[0]}"
    return f"={parts[0]}+{parts[1]}"


def apply_dynamic_totals(workbook, total_fixes=None) -> None:
    fixes = list(total_fixes or DYNAMIC_TOTAL_FIXES)
    for sheet_name, total_cell, col, start_row in fixes:
        resolved_name = _resolve_sheet_name(workbook, sheet_name)
        if not resolved_name:
            continue

        ws = workbook[resolved_name]
        last_row = _find_last_numeric_row(ws, col, start_row)
        if last_row is None or last_row < start_row:
            ws[total_cell].value = f"=SUM({col}{start_row}:{col}{start_row})"
            continue

        ws[total_cell].value = _build_dynamic_sum_formula(
            total_cell=total_cell,
            col=col,
            start_row=start_row,
            end_row=last_row,
         )

def _count_invalid_style_refs(xlsx_path: str) -> int:
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    q = lambda t: f"{{{ns}}}{t}"

    with zipfile.ZipFile(xlsx_path) as z:
        styles_xml = ET.fromstring(z.read("xl/styles.xml"))
        fonts = len(styles_xml.find(q("fonts")).findall(q("font")))
        fills = len(styles_xml.find(q("fills")).findall(q("fill")))
        borders = len(styles_xml.find(q("borders")).findall(q("border")))
        cell_style_xfs = styles_xml.find(q("cellStyleXfs")).findall(q("xf"))
        cell_xfs = styles_xml.find(q("cellXfs")).findall(q("xf"))

        bad = 0
        for xf in cell_xfs:
            font_id = int(xf.attrib.get("fontId", "0"))
            fill_id = int(xf.attrib.get("fillId", "0"))
            border_id = int(xf.attrib.get("borderId", "0"))
            xf_id = int(xf.attrib.get("xfId", "0")) if "xfId" in xf.attrib else None
            if (
                font_id >= fonts
                or fill_id >= fills
                or border_id >= borders
                or (xf_id is not None and xf_id >= len(cell_style_xfs))
            ):
                bad += 1
        return bad


def inspect_new_layout_template(template_xlsx: str) -> dict:
    if not os.path.exists(template_xlsx):
        raise FileNotFoundError(f"Final template workbook not found: {template_xlsx}")

    wb = load_workbook(template_xlsx, read_only=True, data_only=False)
    try:
        sheetnames = wb.sheetnames
        required = ["B3", "B4A Ilum. Pública", "A4 VERDE"]
        missing = [name for name in required if not _sheet_exists_like(sheetnames, name)]
        has_dimensionamento = _sheet_exists_like(sheetnames, "Dimensionamento")
        invalid_style_refs = _count_invalid_style_refs(template_xlsx)
        return {
            "sheetnames": sheetnames,
            "missing_required": missing,
            "has_dimensionamento": has_dimensionamento,
            "invalid_style_refs": invalid_style_refs,
        }
    finally:
        close_fn = getattr(wb, "close", None)
        if callable(close_fn):
            close_fn()


def resolve_default_new_layout_template(base_dir: str, municipio: str) -> str | None:
    template_dir = os.path.join(base_dir, "template")

    # Global one-shot default (shared template for all municipalities/projects)
    # takes precedence over ad-hoc local templates.
    for rel in GLOBAL_TEMPLATE_CANDIDATES:
        p = os.path.abspath(rel)
        if not os.path.exists(p):
            continue
        try:
            info = inspect_new_layout_template(p)
        except Exception:
            continue
        if not info["missing_required"] and info["invalid_style_refs"] == 0:
            return p

    municipio_ascii = "".join(
        ch for ch in unicodedata.normalize("NFKD", municipio) if not unicodedata.combining(ch)
    )
    preferred = [
        os.path.join(template_dir, f"{municipio}_final_template_all_in_one.xlsx"),
        os.path.join(template_dir, f"{municipio_ascii}_final_template_all_in_one.xlsx"),
        os.path.join(template_dir, f"{municipio}_new_layout_template.xlsx"),
        os.path.join(template_dir, f"{municipio_ascii}_new_layout_template.xlsx"),
        os.path.join(template_dir, "final_template_all_in_one.xlsx"),
        os.path.join(template_dir, "new_layout_template.xlsx"),
    ]

    candidates = [p for p in preferred if os.path.exists(p)]
    if os.path.isdir(template_dir):
        xlsx_files = [
            os.path.join(template_dir, name)
            for name in os.listdir(template_dir)
            if name.lower().endswith(".xlsx")
            and not name.lower().endswith("_final_model_filled.xlsx")
            and not name.lower().endswith("_dimensionamento_model_filled.xlsx")
            and not name.lower().endswith("_dimensionamento.xlsx")
        ]
        xlsx_files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        for p in xlsx_files:
            if p not in candidates:
                candidates.append(p)

    inspected: list[tuple[str, dict]] = []
    for p in candidates:
        try:
            info = inspect_new_layout_template(p)
        except Exception:
            continue
        if not info["missing_required"] and info["invalid_style_refs"] == 0:
            inspected.append((p, info))

    # Prefer consolidated templates that already include Dimensionamento.
    for p, info in inspected:
        if info["has_dimensionamento"]:
            return p
    for p, _info in inspected:
        return p

    return None


def disponibilidade_from_tipo(tipo: str) -> int:
    t = (tipo or "").upper()
    if "TRIF" in t:
        return 100
    if "BIF" in t:
        return 50
    if "MONO" in t:
        return 30
    return 0



def ensure_dirs(base_dir: str, create_keys: list[str] | None = None) -> dict:
    paths = {
        "base": base_dir,
        "pdf": os.path.join(base_dir, "pdf"),
        "extracted": os.path.join(base_dir, "extracted"),
        "master": os.path.join(base_dir, "master"),
        "template": os.path.join(base_dir, "template"),
        "dimensionamento": os.path.join(base_dir, "dimensionamento"),
    }
    os.makedirs(paths["base"], exist_ok=True)
    for key in (create_keys or []):
        if key not in paths:
            raise KeyError(f"Unknown output path key: {key}")
        os.makedirs(paths[key], exist_ok=True)
    return paths


def _create_builtin_final_workbook(path: str, month_slots: int = 13) -> None:
    slots = int(month_slots) if isinstance(month_slots, int) and month_slots > 0 else 13
    slots = max(13, slots)

    wb = Workbook()

    ws_b3 = wb.active
    ws_b3.title = "B3"
    b3_headers = [
        "LOCAL",
        "ENDERECO",
        "UC",
        "CLASSIFICACAO",
        "TIPO_FORNECIMENTO",
        "DISPONIBILIDADE",
        "SALDO",
    ]
    for c, header in enumerate(b3_headers, start=1):
        ws_b3.cell(row=4, column=c).value = header
    for i in range(slots):
        ws_b3.cell(row=4, column=8 + i).value = i + 1
    ws_b3.cell(row=4, column=8 + slots).value = "MEDIA_MENSAL"
    ws_b3.cell(row=4, column=9 + slots).value = "PDF_INFO"

    ws_ip = wb.create_sheet("B4A Ilum. Pública")
    ip_headers = [
        "LOCAL",
        "ENDERECO",
        "UC",
        "CLASSIFICACAO",
        "TIPO_FORNECIMENTO",
        "DISPONIBILIDADE",
        "SALDO",
    ]
    for c, header in enumerate(ip_headers, start=1):
        ws_ip.cell(row=4, column=c).value = header
    for i in range(13):
        ws_ip.cell(row=4, column=8 + i).value = i + 1
    ws_ip.cell(row=4, column=21).value = "MEDIA_MENSAL"
    ws_ip.cell(row=4, column=22).value = "PDF_INFO"

    ws_a4 = wb.create_sheet("A4 VERDE")
    a4_headers = {
        1: "LOCAL",
        2: "ENDERECO",
        3: "UC",
        4: "CLASSIFICACAO",
        5: "DEMANDA_CONTRATADA_UNICA_KW",
        6: "DEMANDA_ULTRAPASSADA_KW",
        7: "MAIOR_DEMANDA_MEDIDA_KW",
        47: "MEDIA_CONSUMO_HP",
        61: "MEDIA_CONSUMO_FHP",
        62: "PDF_INFO",
    }
    for c, header in a4_headers.items():
        ws_a4.cell(row=2, column=c).value = header
    for i in range(13):
        ws_a4.cell(row=2, column=8 + i).value = i + 1
        ws_a4.cell(row=2, column=21 + i).value = i + 1
        ws_a4.cell(row=2, column=34 + i).value = i + 1
        ws_a4.cell(row=2, column=48 + i).value = i + 1

    ws_dim = wb.create_sheet("Dimensionamento Prévio UFVs")
    ws_dim.cell(row=1, column=1).value = "Dimensionamento Prévio UFVs"
    ws_dim.cell(row=5, column=3).value = "Consumo B3 medio mensal (kWh)"
    ws_dim.cell(row=13, column=3).value = "Consumo IP medio mensal (kWh)"
    ws_dim.cell(row=24, column=3).value = "Consumo A4 HP medio mensal (kWh)"
    ws_dim.cell(row=32, column=3).value = "Consumo A4 FHP medio mensal (kWh)"

    wb.save(path)


def save_df_csv_tsv(df: pd.DataFrame, csv_path: str, tsv_path: str) -> None:
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_csv(tsv_path, sep="\t", index=False, encoding="utf-8-sig")


def _cell_has_formula(ws, row: int, col: int) -> bool:
    v = ws.cell(row=row, column=col).value
    return isinstance(v, str) and v.startswith("=")


def _clear_cell_preserving_formula(ws, row: int, col: int) -> None:
    if _cell_has_formula(ws, row, col):
        return
    ws.cell(row=row, column=col).value = None


def _write_value_preserving_formula(ws, row: int, col: int, value) -> bool:
    if _cell_has_formula(ws, row, col):
        return False
    ws.cell(row=row, column=col).value = value
    return True


def _copy_formula_from_row(ws, src_row: int, dst_row: int, col: int) -> bool:
    src = ws.cell(row=src_row, column=col).value
    if not (isinstance(src, str) and src.startswith("=")):
        return False
    dst = ws.cell(row=dst_row, column=col)
    if isinstance(dst.value, str) and dst.value.startswith("="):
        return True
    try:
        dst.value = Translator(src, origin=ws.cell(row=src_row, column=col).coordinate).translate_formula(
            dst.coordinate
        )
    except Exception:
        dst.value = src
    return True


def ref_to_date(ref: str):
    # "MM/YYYY" -> Timestamp(YYYY, MM, 1)
    try:
        mm, yyyy = ref.split("/")
        return pd.Timestamp(int(yyyy), int(mm), 1)
    except Exception:
        return pd.NaT



def get_month_refs(master: pd.DataFrame, months_to_use: int = 13) -> list[str]:
    if "referencia" not in master.columns:
        print("WARNING: Column 'referencia' not found in master. No month references available.")
        return []


    # 1) get unique refs
    refs = [
        r for r in master["referencia"].dropna().unique().tolist()
        if isinstance(r, str) and "/" in r
    ]


    # 2) parse and keep only valid refs, then sort chronologically
    parsed_refs = [(r, ref_to_date(r)) for r in refs]
    valid_refs = [(r, d) for r, d in parsed_refs if not pd.isna(d)]
    invalid_count = len(parsed_refs) - len(valid_refs)
    if invalid_count:
        print(f"WARNING: Ignoring {invalid_count} invalid month references.")
    refs = [r for r, _ in sorted(valid_refs, key=lambda x: x[1])]

    # 3) build a contiguous month window ending at the latest detected reference.
    # For template month slots (1..13), keep latest month first.
    # This guarantees fixed-width templates and allows missing months to be zero-filled downstream.
    if months_to_use and valid_refs:
        latest_date = max(d for _, d in valid_refs)
        refs = []
        for i in range(0, months_to_use):
            d = latest_date - pd.DateOffset(months=i)
            refs.append(f"{int(d.month):02d}/{int(d.year)}")
    elif months_to_use and len(refs) > months_to_use:
        refs = refs[-months_to_use:]
        refs = sorted(refs, key=ref_to_date, reverse=True)
    else:
        refs = sorted(refs, key=ref_to_date, reverse=True)

    # 4) warn if too few
    if len(refs) < 12:
        expected = str(months_to_use) if months_to_use else "at least 12"
        print(f"WARNING: Only {len(refs)} month references found. Expected {expected}.")


    # 5) show what we're using
    if refs:
        print(f"Using {len(refs)} months: {refs[0]} -> {refs[-1]} (latest -> oldest)")


    return refs



def validate_master_vs_template(master: pd.DataFrame, template_path: str, sheet: str = "GERAL") -> None:
    if not os.path.exists(template_path):
        print(f"WARNING: Template file not found for validation: {template_path}")
        return


    tmpl = pd.read_excel(template_path, sheet_name=sheet)


    # month columns detected from template
    month_cols = [c for c in tmpl.columns if isinstance(c, str) and "/" in c]


    # same month list used in master logic
    month_refs = get_month_refs(master, months_to_use=MONTHS_TO_USE)
    if not month_refs:
        print("\n=== VALIDATION REPORT ===")
        print("No valid month references found in master. Validation skipped.")
        return


    master_used_total = master[master["referencia"].isin(month_refs)]["kwh_total_te"].fillna(0).sum()
    template_total = tmpl[month_cols].fillna(0).sum().sum()


    print("\n=== VALIDATION REPORT ===")
    print(f"Months detected in master: {master['referencia'].nunique()}")
    print(f"Months used (selected): {month_refs[0]} -> {month_refs[-1]} (count={len(month_refs)})")
    print(f"Template month columns count: {len(month_cols)}")
    print(f"Master total (selected months)= {master_used_total}")
    print(f"Template total                = {template_total}")
    print(f"Difference                    = {master_used_total - template_total}")



def report_uc_month_duplicates(master: pd.DataFrame) -> None:
    required = {"uc", "referencia"}
    if not required.issubset(set(master.columns)):
        print("\n=== DUPLICATE CHECK (UC + Month) ===")
        print("Required columns not found in master. Duplicate check skipped.")
        return


    # Count how many rows exist for each (UC, referencia)
    g = master.groupby(["uc", "referencia"]).size()


    dup_count = int((g > 1).sum())
    max_dup = int(g.max()) if len(g) else 0
    total_pairs = int(len(g))


    print("\n=== DUPLICATE CHECK (UC + Month) ===")
    print(f"Unique UC-month pairs: {total_pairs}")
    print(f"Duplicate UC-month pairs (count>1): {dup_count}")
    print(f"Max rows in a single UC-month: {max_dup}")


    # Show top 10 worst cases (only if duplicates exist)
    if dup_count > 0:
        worst = g[g > 1].sort_values(ascending=False).head(10)
        print("\nTop duplicate UC-month cases (up to 10):")
        for (uc, ref), cnt in worst.items():
            print(f"  UC {uc} | {ref} -> {cnt} rows")



def _first_non_empty(series: pd.Series) -> str:
    for v in series.tolist():
        s = str(v).strip()
        if s and s.lower() not in {"nan", "none"}:
            return s
    return ""



def _normalize_a4_classificacao(raw_value: str) -> str:
    s = _normalize_text(raw_value or "")
    if "a4" in s and "verde" in s:
        return "A4 VERDE"
    if "a4" in s:
        return "A4"
    return "A4"


def _should_fill_zero_months_with_disponibilidade(municipio: str) -> bool:
    # Rule used in final-model templates:
    # when a monthly slot is zero, use disponibilidade equivalent.
    _ = municipio
    return True


def _fill_zero_series_with_disponibilidade(values, disponibilidade: float, months: int) -> list[float]:
    seq = values if isinstance(values, (list, tuple)) else []
    out: list[float] = []
    fill_val = float(disponibilidade or 0.0)
    for i in range(months):
        raw = seq[i] if i < len(seq) else 0.0
        try:
            num = float(raw or 0.0)
        except Exception:
            num = 0.0
        if fill_val > 0 and abs(num) <= 1e-9:
            out.append(fill_val)
        else:
            out.append(num)
    return out


def _order_by_first_page_seen(
    table_df: pd.DataFrame,
    source_df: pd.DataFrame,
    uc_col: str,
) -> pd.DataFrame:
    """
    Keep UC rows ordered by first page appearance in the source extraction.
    Falls back to original table order when page metadata is missing.
    """
    if table_df.empty or source_df.empty or uc_col not in table_df.columns:
        return table_df
    if "uc" not in source_df.columns or "page_first_seen" not in source_df.columns:
        return table_df

    order_map = source_df[["uc", "page_first_seen"]].copy()
    order_map["uc"] = order_map["uc"].astype(str)
    order_map["page_first_seen"] = pd.to_numeric(order_map["page_first_seen"], errors="coerce")
    order_map = (
        order_map.dropna(subset=["page_first_seen"])
        .groupby("uc", as_index=False, sort=False)["page_first_seen"]
        .min()
        .sort_values("page_first_seen", kind="stable")
        .reset_index(drop=True)
    )
    if order_map.empty:
        return table_df

    order_map["_row_order"] = range(len(order_map))
    order_map = order_map.rename(columns={"uc": "_uc_key"})

    out = table_df.copy()
    out["_uc_key"] = out[uc_col].astype(str)
    out = (
        out.merge(order_map[["_uc_key", "_row_order"]], on="_uc_key", how="left")
        .sort_values("_row_order", kind="stable", na_position="last")
        .drop(columns=["_uc_key", "_row_order"])
    )
    return out


def _build_final_category_table(
    master: pd.DataFrame,
    month_refs: list[str],
    category_values: set[str],
    classif_label: str,
    fill_zero_with_disponibilidade: bool = False,
) -> pd.DataFrame:
    m = master[master["categoria"].isin(category_values)].copy()
    if m.empty:
        cols = [
            "LOCAL",
            "ENDERECO",
            "UC",
            "CLASSIFICACAO",
            "TIPO_FORNECIMENTO",
            "DISPONIBILIDADE",
            "SALDO",
            *month_refs,
            "MEDIA_MENSAL",
            "PDF_INFO",
        ]
        return pd.DataFrame(columns=cols)


    m["uc"] = m["uc"].astype(str)


    base = (
        m.groupby("uc", as_index=False)
        .agg(
            nome=("nome", _first_non_empty),
            endereco=("endereco", _first_non_empty),
            tipo_fornecimento=("tipo_fornecimento", _first_non_empty),
            audit_pdf_pages=("audit_pdf_pages", _first_non_empty),
            pdf_source=("pdf_source", _first_non_empty),
        )
        .copy()
    )


    base["DISPONIBILIDADE"] = base["tipo_fornecimento"].apply(disponibilidade_from_tipo).astype(float)
    base["CLASSIFICACAO"] = classif_label
    base["PDF_INFO"] = base["audit_pdf_pages"]
    mask_missing_pdf_info = base["PDF_INFO"].astype(str).str.strip().isin({"", "nan", "None"})
    base.loc[mask_missing_pdf_info, "PDF_INFO"] = base.loc[mask_missing_pdf_info, "pdf_source"]


    pivot = (
        m.pivot_table(
            index="uc",
            columns="referencia",
            values="kwh_total_te",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )


    for ref in month_refs:
        if ref not in pivot.columns:
            pivot[ref] = 0.0
    pivot = pivot[["uc"] + month_refs]


    out = base.merge(pivot, on="uc", how="left")
    for ref in month_refs:
        out[ref] = pd.to_numeric(out[ref], errors="coerce").fillna(0.0)

    if fill_zero_with_disponibilidade and month_refs and (classif_label == "B3" or "IP" in category_values):
        for ref in month_refs:
            out[ref] = out[ref].where(out[ref].abs() > 1e-9, out["DISPONIBILIDADE"])

    out["MEDIA_MENSAL"] = out[month_refs].mean(axis=1) if month_refs else 0.0
    out["SALDO"] = out["MEDIA_MENSAL"] - out["DISPONIBILIDADE"]

    out = out.rename(columns={
        "nome": "LOCAL",
        "endereco": "ENDERECO",
        "uc": "UC",
        "tipo_fornecimento": "TIPO_FORNECIMENTO",
    })


    final_cols = [
        "LOCAL",
        "ENDERECO",
        "UC",
        "CLASSIFICACAO",
        "TIPO_FORNECIMENTO",
        "DISPONIBILIDADE",
        "SALDO",
        *month_refs,
        "MEDIA_MENSAL",
        "PDF_INFO",
    ]
    out = out[final_cols].copy()
    out = _order_by_first_page_seen(out, m, uc_col="UC")
    return out



def _detect_month_slots(ws, header_row: int = 4, start_col: int = 8, max_scan: int = 36) -> int:
    slots = 0
    for c in range(start_col, start_col + max_scan):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, (int, float)):
            slots += 1
        else:
            break
    return slots



def _write_final_table_to_sheet(ws, table_df: pd.DataFrame, month_slots: int, start_row: int = 5) -> None:
    if month_slots <= 0:
        return


    month_cols = [c for c in table_df.columns if isinstance(c, str) and "/" in c][:month_slots]
    end_col = 9 + month_slots  # 1..7 fixed + months + media + pdf


    max_row_to_clear = max(ws.max_row, start_row + len(table_df) + 5)
    for r in range(start_row, max_row_to_clear + 1):
        for c in range(1, end_col + 1):
            _clear_cell_preserving_formula(ws, r, c)

    for idx, (_, row) in enumerate(table_df.iterrows(), start=start_row):
        # Preserve template formulas by copying them from the first data row.
        _copy_formula_from_row(ws, start_row, idx, 6)
        _copy_formula_from_row(ws, start_row, idx, 7)
        _copy_formula_from_row(ws, start_row, idx, 8 + month_slots)

        _write_value_preserving_formula(ws, idx, 1, row["LOCAL"])
        _write_value_preserving_formula(ws, idx, 2, row["ENDERECO"])
        _write_value_preserving_formula(ws, idx, 3, row["UC"])
        _write_value_preserving_formula(ws, idx, 4, row["CLASSIFICACAO"])
        _write_value_preserving_formula(ws, idx, 5, row["TIPO_FORNECIMENTO"])
        _write_value_preserving_formula(
            ws, idx, 6, float(row["DISPONIBILIDADE"]) if pd.notna(row["DISPONIBILIDADE"]) else 0.0
        )
        _write_value_preserving_formula(ws, idx, 7, float(row["SALDO"]) if pd.notna(row["SALDO"]) else 0.0)

        for pos, ref in enumerate(month_cols):
            _write_value_preserving_formula(ws, idx, 8 + pos, float(row[ref]))

        _write_value_preserving_formula(
            ws, idx, 8 + month_slots, float(row["MEDIA_MENSAL"]) if pd.notna(row["MEDIA_MENSAL"]) else 0.0
        )
        _write_value_preserving_formula(ws, idx, 9 + month_slots, row["PDF_INFO"])


def _build_a4_final_table(
    master: pd.DataFrame,
    month_refs: list[str],
    fill_zero_with_disponibilidade: bool = False,
) -> pd.DataFrame:
    a4 = master[master["categoria"].astype(str).str.upper() == "A4"].copy()
    if a4.empty:
        cols = [
            "LOCAL",
            "ENDERECO",
            "UC",
            "CLASSIFICACAO",
            "DEMANDA_CONTRATADA_UNICA_KW",
            "MAIOR_DEMANDA_MEDIDA_KW",
            "DEMANDA_ULTRAPASSADA_KW",
            "demanda_hp",
            "demanda_fhp",
            "consumo_hp",
            "consumo_fhp",
            "PDF_INFO",
        ]
        return pd.DataFrame(columns=cols)

    a4["uc"] = a4["uc"].astype(str)
    if "classificacao_uc" not in a4.columns:
        a4["classificacao_uc"] = ""

    base = (
        a4.groupby("uc", as_index=False)
        .agg(
            nome=("nome", _first_non_empty),
            endereco=("endereco", _first_non_empty),
            classificacao_uc=("classificacao_uc", _first_non_empty),
            tipo_fornecimento=("tipo_fornecimento", _first_non_empty),
            audit_pdf_pages=("audit_pdf_pages", _first_non_empty),
            pdf_source=("pdf_source", _first_non_empty),
        )
        .copy()
    )
    base["CLASSIFICACAO"] = base["classificacao_uc"].map(_normalize_a4_classificacao)
    base["DISPONIBILIDADE"] = base["tipo_fornecimento"].apply(disponibilidade_from_tipo).astype(float)
    base["PDF_INFO"] = base["audit_pdf_pages"]
    missing_pdf = base["PDF_INFO"].astype(str).str.strip().isin({"", "nan", "None"})
    base.loc[missing_pdf, "PDF_INFO"] = base.loc[missing_pdf, "pdf_source"]

    metric_specs = [
        ("demanda_hp_kw", "demanda_hp"),
        ("demanda_fhp_kw", "demanda_fhp"),
        ("consumo_hp_kwh", "consumo_hp"),
        ("consumo_fhp_kwh", "consumo_fhp"),
    ]
    pivots: dict[str, pd.DataFrame] = {}
    for metric_col, _ in metric_specs:
        src = a4[["uc", "referencia", metric_col]].copy()
        src[metric_col] = pd.to_numeric(src[metric_col], errors="coerce").fillna(0.0)
        piv = src.pivot_table(
            index="uc",
            columns="referencia",
            values=metric_col,
            aggfunc="sum",
            fill_value=0.0,
        )
        for ref in month_refs:
            if ref not in piv.columns:
                piv[ref] = 0.0
        pivots[metric_col] = piv[month_refs].copy()
    a4_calc = a4[["uc", "demanda_hp_kw", "demanda_fhp_kw", "dif_demanda", "demanda_contratada_kw"]].copy()
    a4_calc["demanda_hp_kw"] = pd.to_numeric(a4_calc["demanda_hp_kw"], errors="coerce").fillna(0.0)
    a4_calc["demanda_fhp_kw"] = pd.to_numeric(a4_calc["demanda_fhp_kw"], errors="coerce").fillna(0.0)
    a4_calc["dif_demanda"] = pd.to_numeric(a4_calc["dif_demanda"], errors="coerce").fillna(0.0)
    a4_calc["demanda_contratada_kw"] = pd.to_numeric(a4_calc["demanda_contratada_kw"], errors="coerce").fillna(0.0)

    def _max_non_zero(series: pd.Series) -> float:
        s = pd.to_numeric(series, errors="coerce").fillna(0.0)
        s = s[s > 0]
        return float(s.max()) if len(s) else 0.0

    lock_df = (
        a4_calc.groupby("uc", as_index=False)
        .agg(
            demanda_contratada=("demanda_contratada_kw", _max_non_zero),
            demanda_ultrapassada=("dif_demanda", _max_non_zero),
            maior_demanda_hp=("demanda_hp_kw", "max"),
            maior_demanda_fhp=("demanda_fhp_kw", "max"),
        )
        .copy()
    )
    lock_df["DEMANDA_CONTRATADA_UNICA_KW"] = lock_df["demanda_contratada"]
    fallback_demanda = lock_df[["maior_demanda_hp", "maior_demanda_fhp"]].max(axis=1)
    lock_df["DEMANDA_CONTRATADA_UNICA_KW"] = lock_df["DEMANDA_CONTRATADA_UNICA_KW"].where(
        lock_df["DEMANDA_CONTRATADA_UNICA_KW"] > 0,
        fallback_demanda,
    )
    lock_df["DEMANDA_ULTRAPASSADA_KW"] = lock_df["demanda_ultrapassada"]
    fallback_ultrapassada = (fallback_demanda - lock_df["DEMANDA_CONTRATADA_UNICA_KW"]).clip(lower=0.0)
    lock_df["DEMANDA_ULTRAPASSADA_KW"] = lock_df["DEMANDA_ULTRAPASSADA_KW"].where(
        lock_df["DEMANDA_ULTRAPASSADA_KW"] > 0,
        fallback_ultrapassada,
    )
    lock_df["MAIOR_DEMANDA_MEDIDA_KW"] = fallback_demanda
    lock_df = lock_df[
        [
            "uc",
            "DEMANDA_CONTRATADA_UNICA_KW",
            "DEMANDA_ULTRAPASSADA_KW",
            "MAIOR_DEMANDA_MEDIDA_KW",
        ]
    ]
    out = base.merge(lock_df, on="uc", how="left")
    out["DEMANDA_CONTRATADA_UNICA_KW"] = pd.to_numeric(out["DEMANDA_CONTRATADA_UNICA_KW"], errors="coerce").fillna(0.0)
    out["MAIOR_DEMANDA_MEDIDA_KW"] = pd.to_numeric(out["MAIOR_DEMANDA_MEDIDA_KW"], errors="coerce").fillna(0.0)
    out["DEMANDA_ULTRAPASSADA_KW"] = pd.to_numeric(out["DEMANDA_ULTRAPASSADA_KW"], errors="coerce").fillna(0.0)

    out = out.rename(columns={"uc": "UC", "nome": "LOCAL", "endereco": "ENDERECO"})
    out["UC"] = out["UC"].astype(str)

    for metric_col, key in metric_specs:
        s = pivots[metric_col].copy()
        s.index = s.index.astype(str)
        out[key] = out["UC"].map(lambda uc: [float(v) for v in s.loc[uc].tolist()] if uc in s.index else [0.0] * len(month_refs))
        if fill_zero_with_disponibilidade and month_refs:
            out[key] = [
                _fill_zero_series_with_disponibilidade(vals, disp, len(month_refs))
                for vals, disp in zip(out[key].tolist(), out["DISPONIBILIDADE"].tolist())
            ]

    final_cols = [
        "LOCAL",
        "ENDERECO",
        "UC",
        "CLASSIFICACAO",
        "DEMANDA_CONTRATADA_UNICA_KW",
        "MAIOR_DEMANDA_MEDIDA_KW",
        "DEMANDA_ULTRAPASSADA_KW",
        "demanda_hp",
        "demanda_fhp",
        "consumo_hp",
        "consumo_fhp",
        "PDF_INFO",
    ]
    out = out[final_cols].copy()
    out = _order_by_first_page_seen(out, a4, uc_col="UC")
    return out


def _write_a4_table_to_sheet(ws, a4_table: pd.DataFrame, month_refs: list[str], start_row: int = 5) -> None:
    # NEW A4 VERDE layout: keep formula columns intact (F, G, AU, BI).
    clear_ranges = [
        (1, 5),    # A:E
        (8, 20),   # H:T
        (21, 33),  # U:AG
        (34, 45),  # AH:AT (do not clear AU)
        (48, 59),  # AV:BH (do not clear BI)
        (62, 62),  # BJ
    ]

    max_row_to_clear = max(ws.max_row, start_row + len(a4_table) + 20)
    for r in range(start_row, max_row_to_clear + 1):
        for c1, c2 in clear_ranges:
            for c in range(c1, c2 + 1):
                _clear_cell_preserving_formula(ws, r, c)

    if a4_table.empty:
        return

    for idx, (_, row) in enumerate(a4_table.iterrows(), start=start_row):
        # Keep template formulas in calculated columns.
        _copy_formula_from_row(ws, start_row, idx, 6)
        _copy_formula_from_row(ws, start_row, idx, 7)
        has_formula_au = _copy_formula_from_row(ws, start_row, idx, 47)
        has_formula_bi = _copy_formula_from_row(ws, start_row, idx, 61)

        _write_value_preserving_formula(ws, idx, 1, row.get("LOCAL", ""))
        _write_value_preserving_formula(ws, idx, 2, row.get("ENDERECO", ""))
        _write_value_preserving_formula(ws, idx, 3, str(row.get("UC", "") or ""))
        ws.cell(row=idx, column=3).number_format = "@"
        _write_value_preserving_formula(ws, idx, 4, row.get("CLASSIFICACAO", "A4"))
        _write_value_preserving_formula(ws, idx, 5, float(row.get("DEMANDA_CONTRATADA_UNICA_KW", 0.0) or 0.0))

        demanda_hp = row.get("demanda_hp", [0.0] * len(month_refs))
        demanda_fhp = row.get("demanda_fhp", [0.0] * len(month_refs))
        consumo_hp = row.get("consumo_hp", [0.0] * len(month_refs))
        consumo_fhp = row.get("consumo_fhp", [0.0] * len(month_refs))

        for pos in range(min(13, len(month_refs))):
            _write_value_preserving_formula(ws, idx, 8 + pos, float(demanda_hp[pos] or 0.0))    # H..T
            _write_value_preserving_formula(ws, idx, 21 + pos, float(demanda_fhp[pos] or 0.0))  # U..AG
            _write_value_preserving_formula(ws, idx, 34 + pos, float(consumo_hp[pos] or 0.0))   # AH..AT
            _write_value_preserving_formula(ws, idx, 48 + pos, float(consumo_fhp[pos] or 0.0))  # AV..BH

        _write_value_preserving_formula(ws, idx, 6, float(row.get("DEMANDA_ULTRAPASSADA_KW", 0.0) or 0.0))  # F
        _write_value_preserving_formula(ws, idx, 7, float(row.get("MAIOR_DEMANDA_MEDIDA_KW", 0.0) or 0.0))  # G

        # Keep template formulas when available; otherwise generate defaults.
        if not has_formula_au and not _cell_has_formula(ws, idx, 47):
            ws.cell(row=idx, column=47).value = f"=AVERAGE(AH{idx}:AT{idx})"  # AU
        if not has_formula_bi and not _cell_has_formula(ws, idx, 61):
            ws.cell(row=idx, column=61).value = f"=AVERAGE(AV{idx}:BH{idx})"  # BI

        _write_value_preserving_formula(ws, idx, 62, row.get("PDF_INFO", ""))  # BJ


def _write_ip_table_to_sheet(ws, table_df: pd.DataFrame, month_refs: list[str], start_row: int = 5) -> None:
    # B4A Ilum. Publica layout (fixed):
    # A LOCAL, B ENDERECO, C UC, D CLASSIFICACAO, E TIPO
    # F DISPONIBILIDADE (formula), G SALDO (formula)
    # H:T months (13), U MEDIA (formula), V PDF.
    clear_ranges = [
        (1, 5),   # A:E
        (8, 20),  # H:T
        (22, 22), # V
    ]

    max_row_to_clear = max(ws.max_row, start_row + len(table_df) + 20)
    for r in range(start_row, max_row_to_clear + 1):
        for c1, c2 in clear_ranges:
            for c in range(c1, c2 + 1):
                _clear_cell_preserving_formula(ws, r, c)

    if table_df.empty:
        return

    month_cols = [c for c in table_df.columns if isinstance(c, str) and "/" in c][:13]
    for idx, (_, row) in enumerate(table_df.iterrows(), start=start_row):
        has_formula_f = _copy_formula_from_row(ws, start_row, idx, 6)
        has_formula_g = _copy_formula_from_row(ws, start_row, idx, 7)
        has_formula_u = _copy_formula_from_row(ws, start_row, idx, 21)

        _write_value_preserving_formula(ws, idx, 1, row.get("LOCAL", ""))
        _write_value_preserving_formula(ws, idx, 2, row.get("ENDERECO", ""))
        _write_value_preserving_formula(ws, idx, 3, str(row.get("UC", "") or ""))
        ws.cell(row=idx, column=3).number_format = "@"
        _write_value_preserving_formula(ws, idx, 4, row.get("CLASSIFICACAO", "B4 A Ilum, Pública"))
        _write_value_preserving_formula(ws, idx, 5, row.get("TIPO_FORNECIMENTO", ""))

        for pos, ref in enumerate(month_cols):
            _write_value_preserving_formula(ws, idx, 8 + pos, float(row.get(ref, 0.0) or 0.0))  # H:T

        # Preserve template formulas; fallback only when template has none.
        if not has_formula_f and not _cell_has_formula(ws, idx, 6):
            ws.cell(row=idx, column=6).value = f'=IF(E{idx}="TRIFÁSICO",100,IF(E{idx}="BIFÁSICO",50,IF(E{idx}="MONOFÁSICO",30,"")))'
        if not has_formula_g and not _cell_has_formula(ws, idx, 7):
            ws.cell(row=idx, column=7).value = f"=U{idx}-F{idx}"
        if not has_formula_u and not _cell_has_formula(ws, idx, 21):
            ws.cell(row=idx, column=21).value = f"=AVERAGE(H{idx}:T{idx})"

        _write_value_preserving_formula(ws, idx, 22, row.get("PDF_INFO", ""))  # V


def export_to_final_workbook(
    master: pd.DataFrame,
    out_dir: str,
    municipio: str,
    template_xlsx: str | None,
    fill_dimensionamento: bool = True,
) -> str:
    """
    Creates a municipality workbook based on the finalized model and writes:
      - B3
      - B4A Ilum. Publica (from categoria IP)
    This is additive and does not replace the current template pipeline.
    """
    # template_xlsx may be None in one-shot mode.

    paths = ensure_dirs(out_dir, create_keys=["template"])
    month_refs = get_month_refs(master, months_to_use=MONTHS_TO_USE)

    out_xlsx = os.path.join(paths["template"], f"{municipio}_final_model_filled.xlsx")
    if template_xlsx and os.path.exists(template_xlsx):
        if os.path.abspath(template_xlsx) != os.path.abspath(out_xlsx):
            shutil.copy2(template_xlsx, out_xlsx)
        else:
            print("Template path is same as output path; skipping copy step.")
    else:
        if template_xlsx:
            print(
                "WARNING: Final template workbook not found. "
                f"Falling back to built-in one-shot layout: {template_xlsx}"
            )
        _create_builtin_final_workbook(out_xlsx, month_slots=max(len(month_refs), 13))
        print("Using built-in one-shot workbook layout.")

    wb = load_workbook(out_xlsx)
    print("Sheets:", wb.sheetnames)
    print("A4 VERDE exists:", "A4 VERDE" in wb.sheetnames)
    filled_sheets: list[str] = []
    use_palhoca_zero_fill = _should_fill_zero_months_with_disponibilidade(municipio)

    if "B3" in wb.sheetnames:
        ws_b3 = wb["B3"]
        b3_table = _build_final_category_table(
            master,
            month_refs,
            {"B3"},
            "B3",
            fill_zero_with_disponibilidade=use_palhoca_zero_fill,
        )
        b3_slots = _detect_month_slots(ws_b3)
        _write_final_table_to_sheet(ws_b3, b3_table, b3_slots)
        filled_sheets.append("B3")

    if "B4A Ilum. Pública" in wb.sheetnames:
        ws_ip = wb["B4A Ilum. Pública"]
        ip_table = _build_final_category_table(
            master,
            month_refs,
            {"IP"},
            "B4 A Ilum, Pública",
            fill_zero_with_disponibilidade=use_palhoca_zero_fill,
        )
        _write_ip_table_to_sheet(ws_ip, ip_table, month_refs, start_row=5)
        filled_sheets.append("B4A Ilum. Pública")

    if "A4 VERDE" in wb.sheetnames:
        ws_a4 = wb["A4 VERDE"]
        a4_rows = master[master["categoria"].astype(str).str.upper() == "A4"].copy()
        print("A4 rows in master:", len(a4_rows))
        if a4_rows.empty or (
            pd.to_numeric(a4_rows.get("demanda_hp_kw"), errors="coerce").fillna(0.0).abs().sum() == 0
            and pd.to_numeric(a4_rows.get("demanda_fhp_kw"), errors="coerce").fillna(0.0).abs().sum() == 0
        ):
            print("WARNING: A4 historico series are empty. Suggestion: run with --expand-a4-historico.")
        a4_table = _build_a4_final_table(
            master,
            month_refs,
            fill_zero_with_disponibilidade=use_palhoca_zero_fill,
        )
        print("A4 table rows:", len(a4_table))
        _write_a4_table_to_sheet(ws_a4, a4_table, month_refs, start_row=5)
        filled_sheets.append("A4 VERDE")
    else:
        print("INFO: No 'A4 VERDE' sheet found; skipping A4 export.")

    if fill_dimensionamento:
        dim_sheet_name = next(
            (name for name in wb.sheetnames if "DIMENSIONAMENTO" in str(name).upper()),
            None,
        )
        if dim_sheet_name is None:
            print("INFO: No Dimensionamento sheet found in final template; skipping inline mapping.")
        else:
            b3_media, ip_media, a4_hp_media, a4_fhp_media = _compute_dimensionamento_inputs(
                master,
                month_refs,
                fill_zero_with_disponibilidade=use_palhoca_zero_fill,
            )
            ws_dim = wb[dim_sheet_name]
            mapped_values = [
                ("D5", 5, 4, b3_media),
                ("D13", 13, 4, ip_media),
                ("D24", 24, 4, a4_hp_media),
                ("D32", 32, 4, a4_fhp_media),
            ]
            written: list[str] = []
            skipped: list[str] = []
            for label, r, c, val in mapped_values:
                if _write_value_preserving_formula(ws_dim, r, c, val):
                    written.append(f"{label}={val:.3f}")
                else:
                    skipped.append(label)
            filled_sheets.append(dim_sheet_name)
            if written:
                print(f"Filled Dimensionamento inputs in final workbook {', '.join(written)}")
            if skipped:
                print(f"Preserved template formulas in Dimensionamento cells: {', '.join(skipped)}")

    apply_dynamic_totals(wb)

    wb.save(out_xlsx)
    print("\n=== FINAL MODEL EXPORT SAVED ===")
    print(f"Saved: {out_xlsx}")
    if filled_sheets:
        print(f"Filled sheets: {', '.join(filled_sheets)}")
    else:
        print("Filled sheets: none")
    return out_xlsx


def _mean_month_total_from_table(table_df: pd.DataFrame, month_refs: list[str]) -> float:
    if table_df.empty or not month_refs:
        return 0.0
    month_cols = [c for c in month_refs if c in table_df.columns]
    if not month_cols:
        return 0.0
    m = table_df[month_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    return float(m.sum(axis=0).mean())


def _mean_month_total_from_list_col(table_df: pd.DataFrame, series_col: str, month_refs: list[str]) -> float:
    months = len(month_refs)
    if table_df.empty or months == 0 or series_col not in table_df.columns:
        return 0.0
    totals = [0.0] * months
    for vals in table_df[series_col].tolist():
        if not isinstance(vals, (list, tuple)):
            continue
        for i in range(min(months, len(vals))):
            try:
                totals[i] += float(vals[i] or 0.0)
            except Exception:
                continue
    return float(sum(totals) / months) if months else 0.0


def _compute_dimensionamento_inputs(
    master: pd.DataFrame,
    month_refs: list[str],
    fill_zero_with_disponibilidade: bool = False,
) -> tuple[float, float, float, float]:
    b3_table = _build_final_category_table(
        master,
        month_refs,
        {"B3"},
        "B3",
        fill_zero_with_disponibilidade=fill_zero_with_disponibilidade,
    )
    ip_table = _build_final_category_table(
        master,
        month_refs,
        {"IP"},
        "B4 A Ilum, Pública",
        fill_zero_with_disponibilidade=fill_zero_with_disponibilidade,
    )
    a4_table = _build_a4_final_table(
        master,
        month_refs,
        fill_zero_with_disponibilidade=fill_zero_with_disponibilidade,
    )

    b3_media = _mean_month_total_from_table(b3_table, month_refs)
    ip_media = _mean_month_total_from_table(ip_table, month_refs)
    a4_hp_media = _mean_month_total_from_list_col(a4_table, "consumo_hp", month_refs)
    a4_fhp_media = _mean_month_total_from_list_col(a4_table, "consumo_fhp", month_refs)
    return b3_media, ip_media, a4_hp_media, a4_fhp_media


def export_dimensionamento_workbook(master: pd.DataFrame, out_dir: str, municipio: str, template_xlsx: str) -> str:
    """
    Copies a Dimensionamento model workbook and fills the key consumption inputs:
      - D5  : B3 monthly average total consumption
      - D13 : IP monthly average total consumption
      - D24 : A4 HP monthly average total consumption
      - D32 : A4 FHP monthly average total consumption
    """
    if not os.path.exists(template_xlsx):
        raise FileNotFoundError(f"Dimensionamento template workbook not found: {template_xlsx}")

    paths = ensure_dirs(out_dir, create_keys=["dimensionamento"])
    month_refs = get_month_refs(master, months_to_use=MONTHS_TO_USE)
    use_palhoca_zero_fill = _should_fill_zero_months_with_disponibilidade(municipio)
    b3_media, ip_media, a4_hp_media, a4_fhp_media = _compute_dimensionamento_inputs(
        master,
        month_refs,
        fill_zero_with_disponibilidade=use_palhoca_zero_fill,
    )

    out_xlsx = os.path.join(paths["dimensionamento"], f"{municipio}_dimensionamento_model_filled.xlsx")
    shutil.copy2(template_xlsx, out_xlsx)
    wb = load_workbook(out_xlsx)

    print("Dimensionamento template sheets:", wb.sheetnames)
    dim_sheet_name = next(
        (name for name in wb.sheetnames if "DIMENSIONAMENTO" in str(name).upper()),
        None,
    )
    if dim_sheet_name is None:
        print("WARNING: No Dimensionamento sheet found in template. Copied workbook without data mapping.")
        apply_dynamic_totals(wb)
        wb.save(out_xlsx)
        return out_xlsx

    ws = wb[dim_sheet_name]
    mapped_values = [
        ("D5", 5, 4, b3_media),
        ("D13", 13, 4, ip_media),
        ("D24", 24, 4, a4_hp_media),
        ("D32", 32, 4, a4_fhp_media),
    ]
    written: list[str] = []
    skipped: list[str] = []
    for label, r, c, val in mapped_values:
        if _write_value_preserving_formula(ws, r, c, val):
            written.append(f"{label}={val:.3f}")
        else:
            skipped.append(label)

    apply_dynamic_totals(wb)

    wb.save(out_xlsx)
    print("\n=== DIMENSIONAMENTO MODEL EXPORT SAVED ===")
    print(f"Saved: {out_xlsx}")
    if written:
        print(f"Filled inputs {', '.join(written)}")
    if skipped:
        print(f"Preserved template formulas in cells: {', '.join(skipped)}")
    return out_xlsx


def build_master(
    pdfs: list[str],
    out_dir: str,
    municipio: str,
    copy_pdfs: bool,
    expand_a4_historico: bool = False,
    save_intermediate: bool = False,
    discovery_mode: bool = False,
    profile: bool = False,
) -> pd.DataFrame:
    started_at = time.perf_counter() if profile else 0.0
    create_keys = []
    if copy_pdfs:
        create_keys.append("pdf")
    if save_intermediate:
        create_keys.extend(["extracted", "master"])
    paths = ensure_dirs(out_dir, create_keys=create_keys)
    if not pdfs:
        raise ValueError("No PDF files provided.")


    # optionally copy PDFs into municipio/pdf/
    if copy_pdfs:
        for pdf in pdfs:
            dst = os.path.join(paths["pdf"], os.path.basename(pdf))
            if os.path.abspath(pdf) != os.path.abspath(dst):
                shutil.copy2(pdf, dst)


    all_dfs = []


    for pdf_path in pdfs:
        pdf_started_at = time.perf_counter() if profile else 0.0
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF not found: {pdf_path}")


        print(f"\n--- Extracting: {pdf_path}")
        extraction_started_at = time.perf_counter() if profile else 0.0
        df = extract_pdf(
            pdf_path,
            expand_a4_historico=expand_a4_historico,
            discovery_mode=discovery_mode,
        )
        if profile:
            extraction_elapsed = time.perf_counter() - extraction_started_at
            print(f"[profile] extract_pdf: {extraction_elapsed:.2f}s")


        # keep UC as text always
        if "uc" in df.columns:
            df["uc"] = df["uc"].astype(str)


        # audit within this PDF
        audit_keys = ["uc", "referencia"] if {"uc", "referencia"}.issubset(df.columns) else ["uc"]
        audit_df = build_audit_pdf_pages(df, group_cols=audit_keys)
        df = df.merge(audit_df, on=audit_keys, how="left")


        if save_intermediate:
            stem = os.path.splitext(os.path.basename(pdf_path))[0]
            out_csv = os.path.join(paths["extracted"], f"{municipio}_{stem}.csv")
            out_tsv = os.path.join(paths["extracted"], f"{municipio}_{stem}.tsv")
            save_df_csv_tsv(df, out_csv, out_tsv)

            print(f"Saved: {out_csv}")
        print(f"rows= {len(df)} | missing_ref= {int(df['referencia'].isna().sum())} | blank_kwh= {int(df['kwh_total_te'].isna().sum())}")
        if profile:
            pdf_elapsed = time.perf_counter() - pdf_started_at
            print(f"[profile] total_pdf: {pdf_elapsed:.2f}s")


        all_dfs.append(df)


    if not all_dfs:
        print("WARNING: No data extracted from provided PDFs.")
        return pd.DataFrame()


    master = pd.concat(all_dfs, ignore_index=True)


    # rebuild audit across ALL PDFs
    master_audit_keys = ["uc", "referencia"] if {"uc", "referencia"}.issubset(master.columns) else ["uc"]
    master_audit = build_audit_pdf_pages(master, group_cols=master_audit_keys)
    master = (
        master.drop(columns=["audit_pdf_pages"], errors="ignore")
        .merge(master_audit, on=master_audit_keys, how="left")
    )


    if save_intermediate:
        master_csv = os.path.join(paths["master"], f"{municipio}_master.csv")
        master_tsv = os.path.join(paths["master"], f"{municipio}_master.tsv")
        save_df_csv_tsv(master, master_csv, master_tsv)

        print("\n=== MASTER SAVED ===")
        print(f"Saved: {master_csv}")
    else:
        print("\n=== MASTER READY (not saved) ===")
        print("Tip: pass --save-intermediate to also write extracted/master CSV+TSV files.")
    print(f"rows= {len(master)}")
    print("categoria counts:")
    print(master["categoria"].value_counts(dropna=False).to_string())
    if profile:
        elapsed = time.perf_counter() - started_at
        print(f"[profile] build_master_total: {elapsed:.2f}s")


    return master



def build_template_from_master(master: pd.DataFrame, out_dir: str, municipio: str) -> pd.DataFrame:
    paths = ensure_dirs(out_dir, create_keys=["template"])
    month_refs = get_month_refs(master, months_to_use=MONTHS_TO_USE)


    # base identity per UC (take first non-empty values)
    base_cols = ["uc", "nome", "endereco", "categoria", "tipo_fornecimento", "audit_pdf_pages"]
    base = (
        master[base_cols]
        .pipe(_order_by_first_page_seen, master, "uc")
        .drop_duplicates(subset=["uc"], keep="first")
        .copy()
    )


    base["DISPONIBILIDADE"] = base["tipo_fornecimento"].apply(disponibilidade_from_tipo)
    base["SALDO"] = ""  # keep blank


    # monthly matrix: UC x referencia
    m = master.copy()
    m["uc"] = m["uc"].astype(str)


    pivot = (
        m.pivot_table(
            index="uc",
            columns="referencia",
            values="kwh_total_te",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )


    for ref in month_refs:
        if ref not in pivot.columns:
            pivot[ref] = 0.0


    pivot = pivot[["uc"] + month_refs]


    # merge base + months
    out = base.merge(pivot, on="uc", how="left")
    for ref in month_refs:
        out[ref] = out[ref].fillna(0.0)
    out = _order_by_first_page_seen(out, master, uc_col="uc")


    # monthly average (including zeros)
    out["MEDIA MENSAL"] = out[month_refs].mean(axis=1) if month_refs else 0.0


    # final column order
    final_cols = [
        "nome",
        "endereco",
        "uc",
        "categoria",
        "tipo_fornecimento",
        "DISPONIBILIDADE",
        "SALDO",
        "audit_pdf_pages",
    ] + month_refs + ["MEDIA MENSAL"]


    out = out[final_cols].copy()


    # rename headers to match your template wording
    out = out.rename(columns={
        "nome": "Nome",
        "endereco": "Endereço",
        "uc": "UC",
        "categoria": "CLASSIFICAÇÃO (CATEGORIA)",
        "tipo_fornecimento": "TIPO DE FORNECIMENTO",
        "audit_pdf_pages": "PDF + PÁGINAS (audit trail)",
    })


    # save
    out_tsv = os.path.join(paths["template"], f"{municipio}_template_final.tsv")
    out_xlsx = os.path.join(paths["template"], f"{municipio}_template_final.xlsx")


    out.to_csv(out_tsv, sep="\t", index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # All UCs
        out.to_excel(writer, sheet_name="GERAL", index=False)


        # Separate by category
        out[out["CLASSIFICAÇÃO (CATEGORIA)"] == "A4"].to_excel(writer, sheet_name="A4", index=False)
        out[out["CLASSIFICAÇÃO (CATEGORIA)"] == "B3"].to_excel(writer, sheet_name="B3", index=False)
        out[out["CLASSIFICAÇÃO (CATEGORIA)"] == "IP"].to_excel(writer, sheet_name="IP", index=False)


    print("\n=== TEMPLATE SAVED ===")
    print(f"Saved: {out_xlsx}")
    print(f"Saved: {out_tsv}")


    return out



def build_dimensionamento(master: pd.DataFrame, out_dir: str, municipio: str) -> pd.DataFrame:
    paths = ensure_dirs(out_dir, create_keys=["dimensionamento"])
    month_refs = get_month_refs(master, months_to_use=MONTHS_TO_USE)


    # total 12m kWh across all rows (UC-month)
    # safer: sum monthly totals via UC x month pivot to avoid duplicates
    m = master.copy()
    m["uc"] = m["uc"].astype(str)


    pivot_all = m.pivot_table(
        index="uc",
        columns="referencia",
        values="kwh_total_te",
        aggfunc="sum",
        fill_value=0,
    )


    for ref in month_refs:
        if ref not in pivot_all.columns:
            pivot_all[ref] = 0.0


    total_kwh = float(pivot_all[month_refs].sum().sum()) if month_refs else 0.0


    # totals by category
    by_cat = (
        m.pivot_table(
            index="categoria",
            columns="referencia",
            values="kwh_total_te",
            aggfunc="sum",
            fill_value=0,
        )
    )
    for ref in month_refs:
        if ref not in by_cat.columns:
            by_cat[ref] = 0.0


    cat_totals = by_cat[month_refs].sum(axis=1).to_dict() if month_refs else {}


    scenarios = [
        ("Conservador", 1250),
        ("Base", 1350),
        ("Otimista", 1450),
    ]


    total_kwh_label = f"Total {len(month_refs)}m (kWh)"
    rows = []
    for name, prod in scenarios:
        kwp = total_kwh / float(prod) if prod else 0.0
        mwp = kwp / 1000.0
        rows.append({
            "Municipio": municipio,
            "Cenario": name,
            "Produtividade (kWh/kWp.ano)": prod,
            total_kwh_label: total_kwh,
            "Potencia Necessaria (kWp)": kwp,
            "Potencia Necessaria (MWp)": mwp,
            f"Total A4 {len(month_refs)}m (kWh)": float(cat_totals.get("A4", 0.0)),
            f"Total B3 {len(month_refs)}m (kWh)": float(cat_totals.get("B3", 0.0)),
            f"Total IP {len(month_refs)}m (kWh)": float(cat_totals.get("IP", 0.0)),
        })


    dim = pd.DataFrame(rows)


    out_xlsx = os.path.join(paths["dimensionamento"], f"{municipio}_dimensionamento.xlsx")
    out_csv = os.path.join(paths["dimensionamento"], f"{municipio}_dimensionamento.csv")


    dim.to_excel(out_xlsx, index=False)
    dim.to_csv(out_csv, index=False, encoding="utf-8-sig")


    print("\n=== DIMENSIONAMENTO SAVED ===")
    print(f"Saved: {out_xlsx}")
    print(f"Total {len(month_refs)}m (kWh)= {total_kwh:,.3f}".replace(",", "X").replace(".", ",").replace("X", "."))


    return dim



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run the full municipality extraction study from one or more invoice PDFs."
        )
    )
    parser.add_argument("municipio", help="Municipality name used for output folder naming")
    parser.add_argument("pdfs", nargs="+", help="Input PDF file paths")
    parser.add_argument(
        "--final-template",
        dest="final_template_xlsx",
        help="Path to consolidated final workbook template",
    )
    parser.add_argument(
        "--copy-pdfs",
        action="store_true",
        help="Copy input PDFs into municipios/<MUNICIPIO>/pdf",
    )
    parser.add_argument(
        "--expand-a4-historico",
        action=argparse.BooleanOptionalAction,
        default=True,
        help=(
            "Expand historical month rows into the output dataset for A4/B3/IP "
            "(use --no-expand-a4-historico to disable)"
        ),
    )
    parser.add_argument(
        "--save-intermediate",
        action="store_true",
        help="Save extracted per-PDF and master CSV/TSV intermediate files",
    )
    parser.add_argument(
        "--save-dimensionamento-summary",
        action="store_true",
        help="Save standalone dimensionamento summary outputs",
    )
    parser.add_argument(
        "--dimensionamento-template",
        dest="dimensionamento_template_xlsx",
        help="Path to template used for separate dimensionamento workbook export",
    )
    parser.add_argument(
        "--separate-dimensionamento-output",
        action="store_true",
        help="Export an additional standalone dimensionamento workbook",
    )
    parser.add_argument(
        "--profile",
        action="store_true",
        help="Print profiling timings for each stage",
    )
    parser.add_argument(
        "--discovery-mode",
        action="store_true",
        help="Enable discovery logging and write <pdf>_discovery.json",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    municipio = args.municipio.strip()
    pdfs = args.pdfs
    copy_pdfs = args.copy_pdfs
    expand_a4_historico = args.expand_a4_historico
    save_intermediate = args.save_intermediate
    save_dimensionamento_summary = args.save_dimensionamento_summary
    discovery_mode = args.discovery_mode
    profile = args.profile
    final_template_xlsx = args.final_template_xlsx
    dimensionamento_template_xlsx = args.dimensionamento_template_xlsx
    separate_dimensionamento_output = bool(
        args.separate_dimensionamento_output or dimensionamento_template_xlsx
    )

    # create municipality folder under ./municipios/<municipio>/
    base_dir = os.path.join("municipios", municipio)

    if not final_template_xlsx:
        final_template_xlsx = resolve_default_new_layout_template(base_dir, municipio)
        if final_template_xlsx:
            print(f"Using default new-layout template: {final_template_xlsx}")
        else:
            print("No template found. One-shot mode will auto-generate workbook layout.")

    template_info = {
        "sheetnames": [],
        "missing_required": [],
        "has_dimensionamento": True,
        "invalid_style_refs": 0,
    }
    if final_template_xlsx:
        template_info = inspect_new_layout_template(final_template_xlsx)
        if template_info["missing_required"]:
            raise ValueError(
                "Final template is missing required sheets: "
                f"{', '.join(template_info['missing_required'])}. "
                f"Found sheets: {template_info['sheetnames']}"
            )
        if template_info["invalid_style_refs"] > 0:
            raise ValueError(
                "Final template has invalid style references "
                f"({template_info['invalid_style_refs']} broken xf entries). "
                "This causes Excel repair errors in /xl/styles.xml. "
                "Use a clean consolidated template."
            )

    if not dimensionamento_template_xlsx and template_info["has_dimensionamento"] and separate_dimensionamento_output:
        if final_template_xlsx:
            dimensionamento_template_xlsx = final_template_xlsx
            print("Using same consolidated template for dimensionamento export.")

    print(f"\n=== FULL STUDY START: {municipio} ===")
    print(f"Output folder: {base_dir}")
    full_started_at = time.perf_counter() if profile else 0.0

    master = build_master(
        pdfs,
        base_dir,
        municipio,
        copy_pdfs=copy_pdfs,
        expand_a4_historico=expand_a4_historico,
        save_intermediate=save_intermediate,
        discovery_mode=discovery_mode,
        profile=profile,
    )
    if master.empty:
        print("\nNo extracted rows found. Stopping after master build.")
        return

    if save_dimensionamento_summary:
        build_dimensionamento(master, base_dir, municipio)
    report_uc_month_duplicates(master)
    final_workbook_xlsx = export_to_final_workbook(
        master,
        base_dir,
        municipio,
        final_template_xlsx,
        fill_dimensionamento=template_info["has_dimensionamento"],
    )
    if separate_dimensionamento_output:
        dim_template = dimensionamento_template_xlsx or final_workbook_xlsx
        export_dimensionamento_workbook(master, base_dir, municipio, dim_template)

    print(f"\n=== DONE: {municipio} ===")
    if profile:
        full_elapsed = time.perf_counter() - full_started_at
        print(f"[profile] full_study_total: {full_elapsed:.2f}s")



if __name__ == "__main__":
    main()




















